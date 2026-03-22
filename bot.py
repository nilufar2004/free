# -*- coding: utf-8 -*-
import logging
import re
from datetime import datetime, timedelta
import warnings
from io import BytesIO
from collections import defaultdict
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove, KeyboardButton, InputFile
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler
from telegram.request import HTTPXRequest
from db_config import DatabaseConnection
from utils import hash_password, verify_password, validate_phone_number
from export_utils import ExportUtils
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
GROUP_CHAT_ID = int(os.getenv('GROUP_CHAT_ID','0') or 0)

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

logger = logging.getLogger(__name__)

# Silence PTB ConversationHandler per_message warnings (non-fatal)
warnings.filterwarnings(
    "ignore",
    message=".*CallbackQueryHandler.*per_message.*ConversationHandler.*",
    category=UserWarning,
)

# Define conversation states
(
    SELECT_ROLE, REGISTER_FIRSTNAME, REGISTER_LASTNAME, 
    REGISTER_PHONE, REGISTER_PASSWORD, VERIFY_PASSWORD, ADMIN_LOGIN, 
    CASHIER_LOGIN, MAIN_MENU, OPEN_SHIFT_AMOUNT, SELECT_LOCATION, 
    UPLOAD_WORKPLACE_STATUS, UPLOAD_TERMINAL_POWER, UPLOAD_ZERO_REPORT, 
    UPLOAD_OPENING_NOTIFICATION, UPLOAD_RECEIPT_ROLL, SUBMIT_SHIFT_OPENING,
    SELECT_PAYMENT_IMAGE, UPLOAD_PAYMENT_IMAGE,
    REPORT_SALES, REPORT_DEBT_RECEIVED, REPORT_EXPENSES, REPORT_UZCARD,
    REPORT_HUMO, REPORT_UZCARD_REFUND, REPORT_HUMO_REFUND, REPORT_OTHER_PAYMENTS,
    REPORT_DEBT_PAYMENTS, REPORT_DEBT_REFUNDS, SUBMIT_DAILY_REPORT, CLOSE_SHIFT,
    EDIT_REPORT_SELECT, EDIT_REPORT_VALUE,
    # Additional states for admin
    ADMIN_REGISTER_PHONE, ADMIN_REGISTER_PASSWORD, ADMIN_VERIFY_PASSWORD
) = range(36)

class SardobaBot:
    def __init__(self):
        self.db = DatabaseConnection()
        self.db.connect()
        self.export_utils = ExportUtils()
        
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start command handler"""
        # Reset transient states on every /start to avoid stale flows
        context.user_data.pop('admin_reports_range_pending', None)
        context.user_data.pop('admin_reports_range_values', None)
        context.user_data.pop('pending_sverka_key', None)
        context.user_data.pop('pending_sverka_state', None)
        context.user_data.pop('pending_edit_key', None)
        context.user_data.pop('pending_payment_image', None)
        context.user_data['flow'] = None
        try:
            # Best-effort: if MySQL restarted, reconnect automatically.
            self.db._ensure_connection()
        except Exception:
            pass
        user = update.effective_user
        # If user already exists, skip registration flow
        try:
            existing = self.db.fetch_one(
                "SELECT * FROM users WHERE telegram_id = %s AND is_active = TRUE",
                (user.id,)
            )
        except Exception as e:
            logger.exception("start(): DB error")
            if update.message:
                await update.message.reply_text("Serverda xatolik. Keyinroq qayta urinib ko'ring.")
            return ConversationHandler.END
        if not existing:
            approved_req = self.db.fetch_one(
                "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'approved' ORDER BY approved_at DESC LIMIT 1",
                (user.id,)
            )
            if approved_req:
                self.db.execute_query(
                    """
                    INSERT INTO users (telegram_id, first_name, last_name, phone_number, role, password_hash, is_active)
                    VALUES (%s, %s, %s, %s, 'cashier', %s, TRUE)
                    """,
                    (
                        approved_req['telegram_id'],
                        approved_req['first_name'],
                        approved_req['last_name'],
                        approved_req['phone_number'],
                        approved_req.get('password_hash')
                    )
                )
                existing = self.db.fetch_one(
                    "SELECT * FROM users WHERE telegram_id = %s AND is_active = TRUE",
                    (user.id,)
                )
        if existing:
            if existing['role'] == 'admin':
                await update.message.reply_text("Xush kelibsiz!")
                await update.message.reply_text("Administrator menyusi:")
                await self.show_admin_menu(update, context)
                return ConversationHandler.END
            # cashier: ask password each time
            cashier_name = (existing.get('first_name') or user.first_name or "").strip()
            await update.message.reply_text(f"Xush kelibsiz, {cashier_name}!")
            context.user_data['cashier_pending_password'] = True
            await update.message.reply_text("Parolni kiriting:")
            return ConversationHandler.END
        # New users go directly to role selection (no language step)
        context.user_data['language'] = 'uz'
        keyboard = [
            [InlineKeyboardButton("Admin", callback_data='role_admin')],
            [InlineKeyboardButton("Kassir", callback_data='role_cashier')]
        ]
        await update.message.reply_text(
            f"Assalomu alaykum, {user.first_name}!\nRol tanlang:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return SELECT_ROLE

    async def select_role(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle role selection"""
        query = update.callback_query
        await query.answer()
        
        role = query.data.split('_')[1]
        context.user_data['role'] = role

        await query.edit_message_text("Ismingizni kiriting:")
        if role == 'admin':
            return ADMIN_LOGIN
        else:
            return REGISTER_FIRSTNAME

    async def register_firstname(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's first name"""
        context.user_data['first_name'] = update.message.text
        await update.message.reply_text("Familiyangizni kiriting:")
        return REGISTER_LASTNAME

    async def register_lastname(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's last name"""
        context.user_data['last_name'] = update.message.text
        await update.message.reply_text("Telefon raqamingizni kiriting (masalan: +998901234567):")
        return REGISTER_PHONE

    async def register_phone(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's phone number"""
        phone = update.message.text

        if not validate_phone_number(phone):
            await update.message.reply_text(
                "Noto'g'ri telefon raqam formati! Iltimos, qaytadan kiriting (masalan: +998901234567):"
            )
            return REGISTER_PHONE

        context.user_data['phone'] = phone

        # Create or update approval request (password will be set after approval)
        user_data = {
            'telegram_id': update.effective_user.id,
            'first_name': context.user_data.get('first_name', ''),
            'last_name': context.user_data.get('last_name', ''),
            'phone_number': context.user_data.get('phone', ''),
            'role': 'cashier'
        }

        existing_req = self.db.fetch_one(
            "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'pending'",
            (user_data['telegram_id'],)
        )
        if existing_req:
            self.db.execute_query(
                """
                UPDATE approval_requests
                SET first_name=%s, last_name=%s, phone_number=%s
                WHERE telegram_id=%s AND status='pending'
                """,
                (user_data['first_name'], user_data['last_name'], user_data['phone_number'], user_data['telegram_id'])
            )
        else:
            query = """
                INSERT INTO approval_requests (telegram_id, first_name, last_name, phone_number, role)
                VALUES (%(telegram_id)s, %(first_name)s, %(last_name)s, %(phone_number)s, %(role)s)
            """
            self.db.execute_query(query, user_data)

        # Notify admins about new cashier request
        await self.notify_admins_new_request(context, user_data)

        lang = 'uz'
        if lang == 'uz':
            msg = "So'rovingiz administratorga jo'natildi. Tasdiqlanganidan keyin parol o'rnatasiz."
        else:
            msg = "Р В РІР‚в„ўР В Р’В°Р РЋРІвЂљВ¬ Р В Р’В·Р В Р’В°Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР РЋР С“ Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦ Р В Р’В°Р В РўвЂР В РЎВР В РЎвЂР В Р вЂ¦Р В РЎвЂР РЋР С“Р РЋРІР‚С™Р РЋР вЂљР В Р’В°Р РЋРІР‚С™Р В РЎвЂўР РЋР вЂљР РЋРЎвЂњ. Р В РЎСџР В РЎвЂўР РЋР С“Р В Р’В»Р В Р’Вµ Р В РЎвЂўР В РўвЂР В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ Р В Р вЂ Р РЋРІР‚в„– Р РЋРЎвЂњР РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В РЎвЂўР В Р’В»Р РЋР Р‰."

        await update.message.reply_text(msg)
        return ConversationHandler.END

    async def register_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's password"""
        password = update.message.text
        context.user_data['password'] = hash_password(password)
        
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Parolni qaytadan kiriting:"
        else:
            msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°:"
        
        await update.message.reply_text(msg)
        return VERIFY_PASSWORD

    async def verify_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Verify password match"""
        password = update.message.text
        stored_password = context.user_data['password']
        
        if verify_password(stored_password, password):
            # Store user data temporarily for admin approval
            user_data = {
                'telegram_id': update.effective_user.id,
                'first_name': context.user_data['first_name'],
                'last_name': context.user_data['last_name'],
                'phone_number': context.user_data['phone'],
                'role': 'cashier'
            }
            
            # Insert or update approval request
            user_data['password_hash'] = stored_password
            existing_req = self.db.fetch_one(
                "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'pending'",
                (user_data['telegram_id'],)
            )
            if existing_req:
                self.db.execute_query(
                    """
                    UPDATE approval_requests
                    SET first_name=%s, last_name=%s, phone_number=%s, password_hash=%s
                    WHERE telegram_id=%s AND status='pending'
                    """,
                    (
                        user_data['first_name'],
                        user_data['last_name'],
                        user_data['phone_number'],
                        user_data['password_hash'],
                        user_data['telegram_id']
                    )
                )
            else:
                query = """
                    INSERT INTO approval_requests (telegram_id, first_name, last_name, phone_number, role, password_hash)
                    VALUES (%(telegram_id)s, %(first_name)s, %(last_name)s, %(phone_number)s, %(role)s, %(password_hash)s)
                """
                self.db.execute_query(query, user_data)

            # Notify admins about new cashier request
            await self.notify_admins_new_request(context, user_data)
            
            lang = 'uz'
            
            if lang == 'uz':
                msg = "So'rovingiz administratorga jo'natildi. Tasdiqlanganidan keyin botdan foydalanishingiz mumkin."
            else:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†РІР‚С™Р’В¬ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦ Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р Р‹Р РЋРІР‚Сљ. Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњР В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В¶Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В Р РЏ Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°Р В Р’В Р вЂ™Р’В·Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В Р РЏ Р В Р’В Р вЂ™Р’В±Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’В."
            
            await update.message.reply_text(msg)
            return ConversationHandler.END
        else:
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Parollar mos kelmadi! Iltimos, qaytadan kiriting:"
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚В Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚в„–Р В Р Р‹Р Р†Р вЂљРЎв„ў! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°:"
            
            await update.message.reply_text(msg)
            return REGISTER_PASSWORD

    async def admin_login(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Admin login flow"""
        # If user is cashier, require admin promotion
        existing_user = self.db.fetch_one(
            "SELECT * FROM users WHERE telegram_id = %s",
            (update.effective_user.id,)
        )
        if existing_user and existing_user.get('role') == 'cashier':
            await update.message.reply_text("Admin bo'lish uchun mavjud admin tasdiqlashi kerak.")
            return ConversationHandler.END

        # Allow up to 2 admins
        admin_count = self.db.fetch_one("SELECT COUNT(*) as cnt FROM users WHERE role = 'admin' AND is_active = TRUE")
        if admin_count and int(admin_count.get('cnt', 0)) >= 2:
            # If this user is already admin, allow login; otherwise block
            existing_admin = self.db.fetch_one(
                "SELECT * FROM users WHERE telegram_id = %s AND role = 'admin'",
                (update.effective_user.id,)
            )
            if not existing_admin:
                await update.message.reply_text("Adminlar limiti 2 ta. Hozircha yangi admin qo'shib bo'lmaydi.")
                return ConversationHandler.END

        # Check if user is already registered as admin
        query = "SELECT * FROM users WHERE telegram_id = %s AND role = 'admin'"
        result = self.db.fetch_one(query, (update.effective_user.id,))

        if result:
            # Admin already exists
            lang = 'uz'
            
            if lang == 'uz':
                msg = f"Xush kelibsiz, Administrator!"
            else:
                msg = f"Р В Р’В Р Р†Р вЂљРЎСљР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°, Р В Р’В Р РЋРІР‚в„ўР В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™!"
                
            await update.message.reply_text(msg)
            await self.show_admin_menu(update, context)
            return MAIN_MENU
        else:
            # New admin registration
            context.user_data['first_name'] = update.effective_user.first_name
            context.user_data['last_name'] = update.effective_user.last_name
            context.user_data['phone'] = ""  # Will be collected
            
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Telefon raqamingizni kiriting (masalan: +998901234567):"
            else:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†РІР‚С™Р’В¬ Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™ Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’ВµР В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В° (Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™: +998901234567):"
                
            await update.message.reply_text(msg)
            return ADMIN_REGISTER_PHONE

    async def admin_register_phone(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get admin's phone number"""
        phone = update.message.text
        
        if not validate_phone_number(phone):
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Noto'g'ri telefon raqam formati! Iltimos, qaytadan kiriting (masalan: +998901234567):"
            else:
                msg = "Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В° Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’ВµР В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В° (Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™: +998901234567):"
                
            await update.message.reply_text(msg)
            return ADMIN_REGISTER_PHONE
        
        context.user_data['phone'] = phone
        
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Parol kiriting:"
        else:
            msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°:"
        
        await update.message.reply_text(msg)
        return ADMIN_REGISTER_PASSWORD

    async def admin_register_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get admin's password"""
        password = update.message.text
        context.user_data['password'] = hash_password(password)
        
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Parolni qaytadan kiriting:"
        else:
            msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°:"
        
        await update.message.reply_text(msg)
        return ADMIN_VERIFY_PASSWORD

    async def admin_verify_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Verify admin password match"""
        password = update.message.text
        stored_password = context.user_data['password']
        
        if verify_password(stored_password, password):
            # Register admin user
            user_data = {
                'telegram_id': update.effective_user.id,
                'first_name': context.user_data['first_name'],
                'last_name': context.user_data['last_name'],
                'phone_number': context.user_data['phone'],
                'role': 'admin',
                'password_hash': context.user_data['password']
            }
            
            query = """
                INSERT INTO users (telegram_id, first_name, last_name, phone_number, role, password_hash)
                VALUES (%(telegram_id)s, %(first_name)s, %(last_name)s, %(phone_number)s, %(role)s, %(password_hash)s)
            """
            self.db.execute_query(query, user_data)
            
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Siz muvaffaqiyatli ro'yxatdan o'tdingiz, Administrator!"
            else:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р Р‹Р РЋРІР‚СљР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†РІР‚С™Р’В¬Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’ВµР В Р’В Р РЋРІР‚вЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р В Р вЂ°, Р В Р’В Р РЋРІР‚в„ўР В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™!"
                
            await update.message.reply_text(msg)
            await self.show_admin_menu(update, context)
            return MAIN_MENU
        else:
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Parollar mos kelmadi! Iltimos, qaytadan kiriting:"
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚В Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚в„–Р В Р Р‹Р Р†Р вЂљРЎв„ў! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°:"
            
            await update.message.reply_text(msg)
            return ADMIN_REGISTER_PASSWORD

    async def show_admin_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show admin menu"""
        context.user_data['admin_reports_range_pending'] = False
        keyboard = [
            [KeyboardButton("Hisobotlar"), KeyboardButton("Barcha kassirlar")],
            [KeyboardButton("Kassir so'rovlari"), KeyboardButton("Ma'lumotlarni o'zgartirish")],
            [KeyboardButton("Excel/PDF yuklab olish")]
        ]
        menu_text = "Administrator menyusi:"

        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(menu_text, reply_markup=reply_markup)

    async def show_admin_reports_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show admin report period options."""
        keyboard = [
            [KeyboardButton("Kunlik"), KeyboardButton("Haftalik")],
            [KeyboardButton("Oylik"), KeyboardButton("Vaqt oralig'i")],
            [KeyboardButton("Orqaga")],
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text("Qaysi hisobot kerak?", reply_markup=reply_markup)

    async def show_cashier_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show cashier menu"""
        keyboard = [
            [KeyboardButton("Smena ochish"), KeyboardButton("Smena yopish")],
            [KeyboardButton("Sverka"), KeyboardButton("Rasm jo'natish")],
            [KeyboardButton("Hisobotlarni tahrirlash")]
        ]
        menu_text = "Kassir menyusi:"

        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(menu_text, reply_markup=reply_markup)

    async def handle_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle general messages based on user state"""
        text = update.message.text
        user_id = update.effective_user.id

        # If cashier is mid-flow, let ConversationHandler handle and avoid menu spam
        if context.user_data.get('flow') in ['opening', 'sverka', 'closing', 'edit', 'payment_image']:
            if context.user_data.get('flow') == 'payment_image' and context.user_data.get('pending_payment_image'):
                menu_texts = {"Smena ochish", "Smena yopish", "Sverka", "Rasm jo'natish", "Hisobotlarni tahrirlash"}
                if text in menu_texts:
                    context.user_data.pop('pending_payment_image', None)
                    context.user_data['flow'] = None
                    # oqimni tozaladik, endi pastdagi oddiy menyu dispatch ishlasin
                else:
                    await update.message.reply_text("Iltimos, rasm yuboring (foto yoki image fayl).")
                    return
            if context.user_data.get('flow') == 'sverka' and context.user_data.get('pending_sverka_key'):
                key = context.user_data.get('pending_sverka_key')
                handlers = {
                    'sales_amount': self.report_sales,
                    'debt_received': self.report_debt_received,
                    'expenses': self.report_expenses,
                    'uzcard_amount': self.report_uzcard,
                    'humo_amount': self.report_humo,
                    'uzcard_refund': self.report_uzcard_refund,
                    'humo_refund': self.report_humo_refund,
                    'other_payments': self.report_other_payments,
                    'debt_payments': self.report_debt_payments,
                    'debt_refunds': self.report_debt_refunds,
                }
                handler = handlers.get(key)
                if handler:
                    await handler(update, context)
                return
            if context.user_data.get('flow') == 'edit' and context.user_data.get('pending_edit_key'):
                await self.edit_reports_value(update, context)
                return
            # Fallback: agar ConversationHandler state yo'qolgan bo'lsa ham oqim davom etsin
            if context.user_data.get('flow') == 'closing':
                await self.close_shift(update, context)
                return
            if context.user_data.get('flow') == 'opening':
                # Location tanlanganidan keyin summa kiritish bosqichi
                if context.user_data.get('location_id') and not context.user_data.get('opening_stage'):
                    await self.open_shift_amount(update, context)
                    return
                # Rasm bosqichlarida matn yuborilsa tushunarli ogohlantirish beramiz
                await update.message.reply_text("Iltimos, rasm yuboring.")
                return
        
        # Check if user is admin or cashier
        query = "SELECT * FROM users WHERE telegram_id = %s AND is_active = TRUE"
        user = self.db.fetch_one(query, (user_id,))
        
        if not user:
            approved_req = self.db.fetch_one(
                "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'approved' ORDER BY approved_at DESC LIMIT 1",
                (user_id,)
            )
            if approved_req:
                self.db.execute_query(
                    """
                    INSERT INTO users (telegram_id, first_name, last_name, phone_number, role, password_hash, is_active)
                    VALUES (%s, %s, %s, %s, 'cashier', %s, TRUE)
                    """,
                    (
                        approved_req['telegram_id'],
                        approved_req['first_name'],
                        approved_req['last_name'],
                        approved_req['phone_number'],
                        approved_req.get('password_hash')
                    )
                )
                user = self.db.fetch_one(query, (user_id,))
                if user:
                    cashier_name = (user.get('first_name') or update.effective_user.first_name or "").strip()
                    await update.message.reply_text(f"Xush kelibsiz, {cashier_name}!")
                    context.user_data['cashier_pending_password'] = True
                    await update.message.reply_text("Parolni kiriting:")
                    return
            # Check if it's a pending approval request
            query = "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'pending'"
            req = self.db.fetch_one(query, (user_id,))
            if req:
                lang = 'uz'
                
                if lang == 'uz':
                    msg = "Siz hali administrator tomonidan tasdiqlanmadingiz."
                else:
                    msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљР’В°Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В¶Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’В."
                    
                await update.message.reply_text(msg)
                return ConversationHandler.END
            else:
                # First time user, start registration
                await self.start(update, context)
                return SELECT_ROLE
        
        # Determine user role and handle accordingly
        if user['role'] == 'admin':
            await self.handle_admin_command(update, context, user)
        else:
            # If password missing, force set new password
            if not user.get('password_hash'):
                if context.user_data.get('cashier_set_password'):
                    context.user_data['new_password_hash'] = hash_password(text)
                    context.user_data['cashier_set_password'] = False
                    context.user_data['cashier_set_password_confirm'] = True
                    await update.message.reply_text("Parolni qaytadan kiriting:")
                    return
                if context.user_data.get('cashier_set_password_confirm'):
                    if verify_password(context.user_data['new_password_hash'], text):
                        self.db.execute_query(
                            "UPDATE users SET password_hash = %s WHERE telegram_id = %s",
                            (context.user_data['new_password_hash'], user_id)
                        )
                        context.user_data['cashier_set_password_confirm'] = False
                        context.user_data['cashier_authenticated'] = True
                        await update.message.reply_text("Parol o'rnatildi.")
                        await self.show_cashier_menu(update, context)
                    else:
                        context.user_data['cashier_set_password'] = True
                        context.user_data['cashier_set_password_confirm'] = False
                        await update.message.reply_text("Parollar mos kelmadi. Yangi parol kiriting:")
                    return
                context.user_data['cashier_set_password'] = True
                await update.message.reply_text("Parol o'rnatilmagan. Yangi parol kiriting:")
                return

            # Require password on each new /start or session
            if context.user_data.get('cashier_pending_password'):
                menu_texts = {"Smena ochish", "Smena yopish", "Sverka", "Rasm jo'natish", "Hisobotlarni tahrirlash"}
                if text in menu_texts:
                    await update.message.reply_text("Avval parolni kiriting.")
                    return
                if user.get('password_hash') and verify_password(user['password_hash'], text):
                    context.user_data['cashier_pending_password'] = False
                    context.user_data['cashier_authenticated'] = True
                    await self.show_cashier_menu(update, context)
                else:
                    await update.message.reply_text(
                        "Parol noto'g'ri. Qayta kiriting.\n"
                        "Agar parol esdan chiqqan bo'lsa, adminga `reset <telegram_id>` yozdiring."
                    )
                return
            if not context.user_data.get('cashier_authenticated'):
                context.user_data['cashier_pending_password'] = True
                await update.message.reply_text("Parolni kiriting:")
                return
            await self.handle_cashier_command(update, context, user)

    async def handle_image_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle image messages even if ConversationHandler state was lost."""
        flow = context.user_data.get('flow')
        if context.user_data.get('pending_payment_image'):
            await self.upload_payment_image(update, context)
            return

        if flow == 'opening' or context.user_data.get('opening_stage'):
            stage = context.user_data.get('opening_stage')
            if stage == 'workplace_status':
                await self.upload_workplace_status(update, context)
                return
            if stage == 'terminal_power':
                await self.upload_terminal_power(update, context)
                return
            if stage == 'zero_report':
                await self.upload_zero_report(update, context)
                return
            if stage == 'opening_notification':
                await self.upload_opening_notification(update, context)
                return
            if stage == 'receipt_roll':
                await self.upload_receipt_roll(update, context)
                return

    async def handle_admin_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user):
        """Handle admin commands"""
        text = (update.message.text or "").strip()
        lang = 'uz'

        # If we are waiting for a date range input, don't treat admin menu buttons as invalid format.
        # Cancel the pending range when user presses another menu item.
        if context.user_data.get('admin_reports_range_pending'):
            admin_menu_texts = {
                "Hisobotlar",
                "Kunlik",
                "Haftalik",
                "Oylik",
                "Vaqt oralig'i",
                "Barcha kassirlar",
                "Kassir so'rovlari",
                "Ma'lumotlarni o'zgartirish",
                "Excel/PDF yuklab olish",
                "Kunlik hisobot (Excel)",
                "Kunlik hisobot (PDF)",
                "Kassirlar bo'yicha (Excel)",
                "Kassirlar bo'yicha (PDF)",
                "Orqaga",
            }
            if text in admin_menu_texts:
                context.user_data['admin_reports_range_pending'] = False
            else:
                await self.handle_admin_reports_range(update, context)
                return

        # Approve/reject cashier requests by command text
        if text:
            normalized = text.strip()
            lower = normalized.lower()
            if lower.startswith(("approve ", "tasdiq ", "tasdiqlash ", "confirm ", "accept ", "odobrit ", "Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™ ")):
                target_id = self._extract_telegram_id(normalized)
                if target_id:
                    await self.approve_cashier(update, context, target_id)
                else:
                    await update.message.reply_text("ID topilmadi. Masalan: approve 123456789")
                return
            if lower.startswith(("reject ", "rad ", "otkaz ", "otklon ", "deny ", "Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В· ", "Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В¦ ")):
                target_id = self._extract_telegram_id(normalized)
                if target_id:
                    await self.reject_cashier(update, context, target_id)
                else:
                    await update.message.reply_text("ID topilmadi. Masalan: reject 123456789")
                return
            if lower.startswith(("reset ", "parol ", "Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В РЎвЂўР В Р’В» ", "Р РЋР С“Р В Р’В±Р РЋР вЂљР В РЎвЂўР РЋР С“ ")):
                target_id = self._extract_telegram_id(normalized)
                if target_id:
                    await self.reset_cashier_password(update, context, target_id)
                else:
                    await update.message.reply_text("ID topilmadi. Masalan: reset 123456789")
                return
        
        if text == "Hisobotlar":
            await self.show_admin_reports_menu(update, context)
        elif text == "Orqaga":
            await self.show_admin_menu(update, context)
        elif text == "Kunlik":
            await self._ask_report_location(update, context, "daily")
        elif text == "Haftalik":
            await self._ask_report_location(update, context, "weekly")
        elif text == "Oylik":
            await self._ask_report_location(update, context, "monthly")
        elif text == "Vaqt oralig'i":
            await self.send_reports(update, context)
        elif text in ["Barcha kassirlar", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р В РЎвЂњР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“"]:
            # Fetch and send all cashiers
            await self.send_all_cashiers(update, context)
        elif text in ["Kassir so'rovlari", "Р В Р вЂ Р РЋРЎв„ўР Р†Р вЂљР’В¦ Р В Р’В Р Р†Р вЂљРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В "]:
            # Handle approval requests
            await self.handle_approval_requests(update, context)
        elif text in ["Ma'lumotlarni o'zgartirish", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРІР‚СњР Р†Р вЂљРЎв„ў Р В Р’В Р вЂ™Р’ВР В Р’В Р вЂ™Р’В·Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ° Р В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљР’В Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚в„–"]:
            # Handle data modification
            await self.modify_user_data(update, context)
        elif text in ["Excel/PDF yuklab olish", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р СћРЎвЂ™ Р В Р’В Р В Р вЂ№Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ° Excel/PDF"]:
            # Handle export to Excel/PDF
            await self.export_data(update, context)
        elif text in ["Kunlik hisobot (Excel)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (Excel)", "Kunlik hisobot (PDF)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (PDF)", "Kassirlar bo'yicha (Excel)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (Excel)", "Kassirlar bo'yicha (PDF)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (PDF)", "Orqaga", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎСљР Р†РІР‚С›РЎС› Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚В"]:
            # Handle export choice
            await self.handle_export_choice(update, context)
        else:
            if lang == 'uz':
                msg = "Iltimos, menyudan birini tanlang."
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В· Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В  Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р В РІР‚в„–."
                
            await update.message.reply_text(msg)
            await self.show_admin_menu(update, context)

    async def handle_cashier_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user):
        """Handle cashier commands"""
        text = update.message.text
        lang = 'uz'

        # Let ConversationHandler handle these menu actions

        # Majburiy jarayonlar ishlayotgan paytda noto'g'ri bosqichga o'tishni bloklaymiz.
        # payment_image holatida esa menyu tugmalari bosilsa oqimni tozalab davom etishga ruxsat beramiz.
        active_flow = context.user_data.get('flow')
        if active_flow in ['opening', 'sverka', 'closing', 'payment_image']:
            if active_flow == 'payment_image':
                menu_texts = {"Smena ochish", "Smena yopish", "Sverka", "Rasm jo'natish", "Hisobotlarni tahrirlash"}
                if text in menu_texts:
                    context.user_data.pop('pending_payment_image', None)
                    context.user_data['flow'] = None
                else:
                    await update.message.reply_text("Iltimos, rasm yuboring yoki menyudan tugma tanlang.")
                    return
            else:
                await update.message.reply_text("Jarayon davom etmoqda. Iltimos, avval joriy bosqichni yakunlang.")
                return
        
        if text == "Smena ochish":
            # Start shift opening process
            await self.start_shift_opening(update, context)
        elif text == "Smena yopish":
            # Start shift closing process
            await self.start_shift_closing(update, context)
        elif text == "Sverka":
            # Start daily reporting process
            await self.start_daily_reporting(update, context)
        elif text == "Rasm jo'natish":
            # Handle payment image upload (Uzcard/Humo)
            await self.start_payment_image_upload(update, context)
        elif text == "Hisobotlarni tahrirlash":
            # Handle report editing
            await self.edit_reports(update, context)
        else:
            if lang == 'uz':
                msg = f"Xush kelibsiz, {user['first_name']}! Iltimos, menyudan birini tanlang."
            else:
                msg = f"Р В Р’В Р Р†Р вЂљРЎСљР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°, {user['first_name']}! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В· Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В  Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р В РІР‚в„–."
                
            await update.message.reply_text(msg)
            await self.show_cashier_menu(update, context)

    async def _ensure_cashier_authenticated(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
        """Kassir uchun parol o'rnatish/kirishni majburiy tekshiradi."""
        tg_id = update.effective_user.id
        user = self.db.fetch_one(
            "SELECT * FROM users WHERE telegram_id=%s AND role='cashier' AND is_active=TRUE",
            (tg_id,)
        )
        if not user:
            await update.message.reply_text("Kassir topilmadi. /start bosing.")
            return False

        # Tasdiqlangan, lekin hali parol o'rnatilmagan kassir
        if not user.get('password_hash'):
            context.user_data['cashier_set_password'] = True
            context.user_data['cashier_set_password_confirm'] = False
            context.user_data['cashier_pending_password'] = False
            context.user_data['cashier_authenticated'] = False
            await update.message.reply_text("Avval parol o'rnating. Yangi parol kiriting:")
            return False

        # Sessiya uchun parol kiritilmagan bo'lsa
        if not context.user_data.get('cashier_authenticated'):
            context.user_data['cashier_pending_password'] = True
            await update.message.reply_text("Parolni kiriting:")
            return False

        return True

    async def start_shift_opening(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start the shift opening process"""
        if not await self._ensure_cashier_authenticated(update, context):
            return MAIN_MENU
        context.user_data.pop('blocked_media_group_id', None)

        lang = 'uz'
        user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if user_row:
            today_shift = self._today_shift_for_user(user_row['id'])
            if today_shift:
                if bool(today_shift.get('is_open')):
                    await update.message.reply_text("Sizda bugungi ochiq smena bor. Avval o'sha smenani yoping.")
                else:
                    await update.message.reply_text(
                        "Siz bugungi smenani allaqachon yopgansiz. Bir kunda faqat 1 ta smena ochiladi."
                    )
                return MAIN_MENU

            active_shift = self.db.fetch_one(
                "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
                (user_row['id'],)
            )
            if active_shift:
                if lang == 'uz':
                    msg = "Sizda ochiq smena bor. Avval yopishingiz kerak."
                else:
                    msg = "Р В Р’В Р В РІвЂљВ¬ Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњ Р В Р Р‹Р РЋРІР‚СљР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ° Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В Р РЏ Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°. Р В Р’В Р В Р вЂ№Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’В° Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚СњР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљР’В."
                await update.message.reply_text(msg)
                return MAIN_MENU
        
        context.user_data['workplace_status_uploaded_ids'] = []
        context.user_data['opening_stage'] = None
        await self.show_location_selection(update, context)
        context.user_data['flow'] = 'opening'
        return SELECT_LOCATION

    async def open_shift_amount(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get the opening amount for the shift"""
        try:
            raw = update.message.text
            # Allow inputs like "12 300", "12,300", "12330 so'm"
            digits = ''.join(ch for ch in raw if ch.isdigit() or ch in ['.', ','])
            amount = float(digits.replace(',', '')) if digits else float(raw)
            context.user_data['opening_amount'] = amount
            context.user_data['opening_amount_time'] = self._format_telegram_time(getattr(update.message, "date", None))
            
            # Create shift now that we have location + amount
            location_id = context.user_data.get('location_id')
            if not location_id:
                await update.message.reply_text("Avval filialni tanlang.")
                await self.show_location_selection(update, context)
                return SELECT_LOCATION

            try:
                user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
                if user_row:
                    today_shift = self._today_shift_for_user(user_row['id'])
                    if today_shift:
                        if bool(today_shift.get('is_open')):
                            await update.message.reply_text("Sizda bugungi ochiq smena bor. Avval yoping.")
                        else:
                            await update.message.reply_text(
                                "Siz bugungi smenani allaqachon yopgansiz. Qayta ochib bo'lmaydi."
                            )
                        await self.show_cashier_menu(update, context)
                        context.user_data['flow'] = None
                        return MAIN_MENU

                    self.db.execute_query(
                        """
                        INSERT INTO shifts (user_id, location_id, opening_amount, is_open)
                        VALUES (%s, %s, %s, TRUE)
                        """,
                        (user_row['id'], location_id, amount)
                    )
                    shift = self.db.fetch_one(
                        "SELECT id FROM shifts WHERE user_id=%s ORDER BY opened_at DESC LIMIT 1",
                        (user_row['id'],)
                    )
                    if shift:
                        context.user_data['current_shift_id'] = shift['id']
                        context.user_data['workplace_status_uploaded_ids'] = []

                # Notify group about shift opening
                loc = self.db.fetch_one("SELECT name FROM locations WHERE id = %s", (location_id,))
                loc_name = loc['name'] if loc else str(location_id)
                await self._send_group_message(
                    context,
                    f"Smena ochildi: {update.effective_user.first_name} {update.effective_user.last_name or ''}\n"
                    f"Filial: {loc_name}\n"
                    f"Ochish summasi: {amount}\n"
                    f"Vaqt: {context.user_data.get('opening_amount_time','')}"
                )
            except Exception:
                context.user_data['flow'] = None
                await update.message.reply_text("Xatolik: smena ma'lumotlarini saqlab bo'lmadi. Qayta urinib ko'ring.")
                await self.show_cashier_menu(update, context)
                return MAIN_MENU

            lang = 'uz'
            if lang == 'uz':
                await update.message.reply_text("Summa tasdiqlandi.")
                msg = "Ish joyingiz tayyorligini tasdiqlang. Ish stolingizni rasmga olib yuboring. (2 ta rasm)"
            else:
                await update.message.reply_text("Р В Р Р‹Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В° Р В РЎвЂ”Р В РЎвЂўР В РўвЂР РЋРІР‚С™Р В Р вЂ Р В Р’ВµР РЋР вЂљР В Р’В¶Р В РўвЂР В Р’ВµР В Р вЂ¦Р В Р’В°.")
                msg = "Р В РЎСџР В РЎвЂўР В РўвЂР РЋРІР‚С™Р В Р вЂ Р В Р’ВµР РЋР вЂљР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ“Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂўР В Р вЂ Р В Р вЂ¦Р В РЎвЂўР РЋР С“Р РЋРІР‚С™Р РЋР Р‰ Р РЋР вЂљР В Р’В°Р В Р’В±Р В РЎвЂўР РЋРІР‚РЋР В Р’ВµР В РЎвЂ“Р В РЎвЂў Р В РЎВР В Р’ВµР РЋР С“Р РЋРІР‚С™Р В Р’В°. Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р РЋР Р‰Р РЋРІР‚С™Р В Р’Вµ Р РЋРІР‚С›Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў Р РЋР вЂљР В Р’В°Р В Р’В±Р В РЎвЂўР РЋРІР‚РЋР В Р’ВµР В РЎвЂ“Р В РЎвЂў Р РЋР С“Р РЋРІР‚С™Р В РЎвЂўР В Р’В»Р В Р’В°. (2 Р РЋРІР‚С›Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў)"
            await update.message.reply_text(msg)
            context.user_data['opening_stage'] = 'workplace_status'
            return UPLOAD_WORKPLACE_STATUS
        except ValueError:
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Iltimos, to'g'ri miqdor kiriting."
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р В РІР‚В Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°Р В Р’В Р В РІР‚В¦Р В Р Р‹Р РЋРІР‚СљР В Р Р‹Р В РІР‚в„– Р В Р Р‹Р В РЎвЂњР В Р Р‹Р РЋРІР‚СљР В Р’В Р РЋР’ВР В Р’В Р РЋР’ВР В Р Р‹Р РЋРІР‚Сљ."
                
            await update.message.reply_text(msg)
            return OPEN_SHIFT_AMOUNT

    async def show_location_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show location selection for the shift"""
        # Get all locations from the database
        locations = self.db.fetch_all("SELECT * FROM locations WHERE is_active = TRUE")
        
        keyboard = []
        for loc in locations:
            keyboard.append([InlineKeyboardButton(loc['name'], callback_data=f"loc_{loc['id']}")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        msg = "Filialni tanlang:"
        await update.message.reply_text(msg, reply_markup=reply_markup)

    async def ask_select_location_again(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Remind user to select location before entering amount"""
        msg = "Avval filialni tanlang."
        await update.message.reply_text(msg)
        await self.show_location_selection(update, context)

    async def select_location(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle location selection"""
        query = update.callback_query
        if not query or not query.data:
            return SELECT_LOCATION

        await query.answer()

        try:
            location_id = int(query.data.split('_')[1])
        except Exception:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Filialni qaytadan tanlang.")
            await self.show_location_selection(update, context)
            return SELECT_LOCATION

        context.user_data['location_id'] = location_id
        context.user_data['flow'] = 'opening'

        msg = "Smenani ochish summasini kiriting (faqat summa):"
        try:
            await query.edit_message_text(msg)
        except Exception:
            # Fallback: agar edit ishlamasa ham keyingi bosqichga o'tkazamiz
            await context.bot.send_message(chat_id=update.effective_chat.id, text=msg)
        return OPEN_SHIFT_AMOUNT
    async def upload_workplace_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload workplace status image (2 ta rasm majburiy)."""
        if self._is_blocked_media_group(update, context):
            return UPLOAD_WORKPLACE_STATUS

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            return UPLOAD_WORKPLACE_STATUS

        shift_id = context.user_data.get('current_shift_id')
        if not shift_id:
            user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id=%s", (update.effective_user.id,))
            if user_row:
                active_shift = self.db.fetch_one(
                    "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
                    (user_row['id'],)
                )
                if active_shift:
                    shift_id = active_shift['id']
                    context.user_data['current_shift_id'] = shift_id

        if not shift_id:
            context.user_data['flow'] = None
            await update.message.reply_text("Ochiq smena topilmadi. Avval smena oching.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU

        uploaded_ids = context.user_data.get('workplace_status_uploaded_ids')
        if not isinstance(uploaded_ids, list):
            uploaded_ids = []

        uploaded_ids.append(file_id)
        self._save_shift_image(shift_id, 'workplace_status', file_id)
        await self._send_group_shift_photo(
            context, shift_id, file_id, "Ish joyi holati rasmi", event_time=getattr(update.message, "date", None)
        )
        context.user_data['workplace_status_uploaded_ids'] = uploaded_ids

        db_count = self._count_shift_images(shift_id, 'workplace_status')
        count = max(len(uploaded_ids), db_count)

        if count < 2:
            await update.message.reply_text("Rasm qabul qilindi (1/2). Yana bitta rasm yuboring.")
            context.user_data['opening_stage'] = 'workplace_status'
            return UPLOAD_WORKPLACE_STATUS

        await update.message.reply_text("Rasmlar qabul qilindi (2/2).")
        await update.message.reply_text(
            "Terminallar va ratsiyalar quvvatini tekshiring va ularning quvvatlanish jarayonini rasmga oling."
        )
        self._block_current_media_group(update, context)
        context.user_data['opening_stage'] = 'terminal_power'
        return UPLOAD_TERMINAL_POWER

    async def upload_terminal_power(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload terminal power image."""
        if self._is_blocked_media_group(update, context):
            return UPLOAD_TERMINAL_POWER

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'terminal_power'
            return UPLOAD_TERMINAL_POWER

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            self._save_shift_image(shift_id, 'terminal_power', file_id)
            await self._send_group_shift_photo(
                context, shift_id, file_id, "Terminal/ratsiya quvvat holati", event_time=getattr(update.message, "date", None)
            )

        await update.message.reply_text("Rasm qabul qilindi.")
        await update.message.reply_text("Uzcard va Humo kartalaridagi nol hisobotni chiqaring va rasmga oling.")
        self._block_current_media_group(update, context)
        context.user_data['opening_stage'] = 'zero_report'
        return UPLOAD_ZERO_REPORT

    async def upload_zero_report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload zero report image."""
        if self._is_blocked_media_group(update, context):
            return UPLOAD_ZERO_REPORT

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'zero_report'
            return UPLOAD_ZERO_REPORT

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            self._save_shift_image(shift_id, 'zero_report', file_id)
            await self._send_group_shift_photo(
                context, shift_id, file_id, "Uzcard/Humo nol hisobot", event_time=getattr(update.message, "date", None)
            )

        await update.message.reply_text("Rasm qabul qilindi.")
        await update.message.reply_text("Iiko va soliq check tizimlarida smenani oching. Ochilganlik haqidagi bildirishnomani rasmga oling.")
        self._block_current_media_group(update, context)
        context.user_data['opening_stage'] = 'opening_notification'
        return UPLOAD_OPENING_NOTIFICATION

    async def upload_opening_notification(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload opening notification image."""
        if self._is_blocked_media_group(update, context):
            return UPLOAD_OPENING_NOTIFICATION

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'opening_notification'
            return UPLOAD_OPENING_NOTIFICATION

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            self._save_shift_image(shift_id, 'opening_notification', file_id)
            await self._send_group_shift_photo(
                context, shift_id, file_id, "iiko/soliq ochilish bildirishnomasi", event_time=getattr(update.message, "date", None)
            )

        await update.message.reply_text("Rasm qabul qilindi.")
        await update.message.reply_text("Zaxira chek lentalari mavjudligini rasm bilan jo'nating.")
        self._block_current_media_group(update, context)
        context.user_data['opening_stage'] = 'receipt_roll'
        return UPLOAD_RECEIPT_ROLL

    async def upload_receipt_roll(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload receipt roll image and finish shift opening flow."""
        if self._is_blocked_media_group(update, context):
            return UPLOAD_RECEIPT_ROLL

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'receipt_roll'
            return UPLOAD_RECEIPT_ROLL

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            self._save_shift_image(shift_id, 'receipt_roll', file_id)
            await self._send_group_shift_photo(
                context, shift_id, file_id, "Zaxira chek lenta rasmi", event_time=getattr(update.message, "date", None)
            )

        await update.message.reply_text("Rasm qabul qilindi.")
        await update.message.reply_text("Smena muvaffaqiyatli ochildi! Endi sverka jarayonini boshlang.")
        self._block_current_media_group(update, context)
        await self.show_cashier_menu(update, context)
        context.user_data['flow'] = None
        context.user_data.pop('opening_stage', None)
        return MAIN_MENU

    async def start_daily_reporting(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start the daily reporting process"""
        if not await self._ensure_cashier_authenticated(update, context):
            return MAIN_MENU

        # Check if there's an active shift
        user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await update.message.reply_text("Foydalanuvchi topilmadi.")
            return MAIN_MENU
        active_shift = self.db.fetch_one("SELECT * FROM shifts WHERE user_id=%s AND is_open=TRUE", (user_row['id'],))
        
        if not active_shift:
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Avval smena ochishingiz kerak."
            else:
                msg = "Р В Р’В Р В Р вЂ№Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’В° Р В Р’В Р В РІР‚В¦Р В Р Р‹Р РЋРІР‚СљР В Р’В Р вЂ™Р’В¶Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ° Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р РЋРІР‚Сљ."
                
            await update.message.reply_text(msg)
            return MAIN_MENU
        
        context.user_data['current_shift_id'] = active_shift['id']
        required_opening = [
            ('workplace_status', "Ish joyi holati rasmi", 2),
            ('terminal_power', "Terminal/ratsiya quvvati rasmi", 1),
            ('zero_report', "Uzcard/Humo nol hisobot rasmi", 1),
            ('opening_notification', "iiko/soliq smena ochilganlik rasmi", 1),
            ('receipt_roll', "Zaxira chek lenta rasmi", 1),
        ]
        missing = []
        for image_type, label, required_count in required_opening:
            current_count = self._count_shift_images(active_shift['id'], image_type)
            if current_count < required_count:
                remaining = required_count - current_count
                if required_count == 1:
                    missing.append(f"- {label}")
                else:
                    missing.append(f"- {label} ({remaining} ta qolgan)")
        if missing:
            msg = (
                "Smena ochish bosqichidagi rasmlar to'liq emas.\n"
                "Quyidagilar yetishmayapti:\n"
                + "\n".join(missing)
            )
            await self.show_opening_requirements_menu(update, context, int(active_shift['id']), note=msg)
            return MAIN_MENU

        # Init sverka status and show interactive checklist
        context.user_data['flow'] = 'sverka'
        context.user_data['sverka_status'] = {key: False for key, *_ in self._sverka_config()}
        self._init_sverka_status(context)
        await self.show_sverka_menu(update, context)
        return SUBMIT_DAILY_REPORT

    async def report_sales(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get sales amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['sales_amount'] = amount
            self._mark_sverka_done(context, 'sales_amount')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_SALES

    async def report_debt_received(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get received debts amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['debt_received'] = amount
            self._mark_sverka_done(context, 'debt_received')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_DEBT_RECEIVED

    async def report_expenses(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get expenses amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['expenses'] = amount
            self._mark_sverka_done(context, 'expenses')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_EXPENSES

    async def report_uzcard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Uzcard amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['uzcard_amount'] = amount
            self._mark_sverka_done(context, 'uzcard_amount')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_UZCARD

    async def report_humo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Humo amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['humo_amount'] = amount
            self._mark_sverka_done(context, 'humo_amount')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_HUMO

    async def report_uzcard_refund(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Uzcard refund amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['uzcard_refund'] = amount
            self._mark_sverka_done(context, 'uzcard_refund')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_UZCARD_REFUND

    async def report_humo_refund(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Humo refund amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['humo_refund'] = amount
            self._mark_sverka_done(context, 'humo_refund')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_HUMO_REFUND

    async def report_other_payments(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get other payments amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['other_payments'] = amount
            self._mark_sverka_done(context, 'other_payments')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_OTHER_PAYMENTS

    async def report_debt_payments(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get debt payments amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['debt_payments'] = amount
            self._mark_sverka_done(context, 'debt_payments')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_DEBT_PAYMENTS

    async def report_debt_refunds(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get debt refunds amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['debt_refunds'] = amount
            self._mark_sverka_done(context, 'debt_refunds')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_DEBT_REFUNDS
    async def save_daily_report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Save the daily report to the database"""
        shift_id = context.user_data['current_shift_id']
        
        report_data = {
            'shift_id': shift_id,
            'report_type': 'daily_report',
            'sales_amount': context.user_data.get('sales_amount', 0),
            'debt_received': context.user_data.get('debt_received', 0),
            'expenses': context.user_data.get('expenses', 0),
            'uzcard_amount': context.user_data.get('uzcard_amount', 0),
            'humo_amount': context.user_data.get('humo_amount', 0),
            'uzcard_refund': context.user_data.get('uzcard_refund', 0),
            'humo_refund': context.user_data.get('humo_refund', 0),
            'other_payments': context.user_data.get('other_payments', 0),
            'debt_payments': context.user_data.get('debt_payments', 0),
            'debt_refunds': context.user_data.get('debt_refunds', 0)
        }
        
        query = """
            INSERT INTO reports (
                shift_id, report_type, sales_amount, debt_received, expenses,
                uzcard_amount, humo_amount, uzcard_refund, humo_refund,
                other_payments, debt_payments, debt_refunds
            ) VALUES (
                %(shift_id)s, %(report_type)s, %(sales_amount)s, %(debt_received)s, %(expenses)s,
                %(uzcard_amount)s, %(humo_amount)s, %(uzcard_refund)s, %(humo_refund)s,
                %(other_payments)s, %(debt_payments)s, %(debt_refunds)s
            )
        """
        self.db.execute_query(query, report_data)
        # Muhim: guruhga bitta yakuniy fayl faqat smena yopilganda yuboriladi.
        # Shu sababli bu yerda alohida sverka fayl yubormaymiz.

    async def start_shift_closing(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start the shift closing process"""
        if not await self._ensure_cashier_authenticated(update, context):
            return MAIN_MENU

        lang = 'uz'
        user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await update.message.reply_text("Foydalanuvchi topilmadi.")
            return MAIN_MENU
        active_shift = self.db.fetch_one(
            "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
            (user_row['id'],)
        )
        if not active_shift:
            if lang == 'uz':
                msg = "Ochiq smena yo'q."
            else:
                msg = "Р В РЎСљР В Р’ВµР РЋРІР‚С™ Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂќР РЋР вЂљР РЋРІР‚в„–Р РЋРІР‚С™Р В РЎвЂўР В РІвЂћвЂ“ Р РЋР С“Р В РЎВР В Р’ВµР В Р вЂ¦Р РЋРІР‚в„–."
            await update.message.reply_text(msg)
            return MAIN_MENU
        # Require sverka before closing
        report = self.db.fetch_one(
            "SELECT id FROM reports WHERE shift_id=%s AND report_type='daily_report' ORDER BY id DESC LIMIT 1",
            (active_shift['id'],)
        )
        if not report:
            if lang == 'uz':
                msg = "Avval sverka tugating. Smenani yopib bo'lmaydi."
            else:
                msg = "Р В Р Р‹Р В Р вЂ¦Р В Р’В°Р РЋРІР‚РЋР В Р’В°Р В Р’В»Р В Р’В° Р В Р’В·Р В Р’В°Р В Р вЂ Р В Р’ВµР РЋР вЂљР РЋРІвЂљВ¬Р В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р В Р вЂ Р В Р’ВµР РЋР вЂљР В РЎвЂќР РЋРЎвЂњ. Р В РЎСљР В Р’ВµР В Р’В»Р РЋР Р‰Р В Р’В·Р РЋР РЏ Р В Р’В·Р В Р’В°Р В РЎвЂќР РЋР вЂљР РЋРІР‚в„–Р РЋРІР‚С™Р РЋР Р‰ Р РЋР С“Р В РЎВР В Р’ВµР В Р вЂ¦Р РЋРЎвЂњ."
            await update.message.reply_text(msg)
            return MAIN_MENU
        context.user_data['current_shift_id'] = active_shift['id']

        # Smena yopishdan oldin Uzcard va Humo rasmlari majburiy
        uzcard_img = self._count_shift_images(active_shift['id'], 'uzcard_payment')
        humo_img = self._count_shift_images(active_shift['id'], 'humo_payment')
        if uzcard_img < 1 or humo_img < 1:
            missing = []
            if uzcard_img < 1:
                missing.append("Uzcard rasmi")
            if humo_img < 1:
                missing.append("Humo rasmi")
            await update.message.reply_text(
                "Smenani yopishdan oldin quyidagilar majburiy:\n- "
                + "\n- ".join(missing)
                + "\n\n`Rasm jo'natish` tugmasini bosib, Uzcard va Humo rasmlarini yuboring."
            )
            return MAIN_MENU

        if lang == 'uz':
            msg = "Smenani yopish uchun yakuniy summani kiriting:"
        else:
            msg = "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂР РЋРІР‚С™Р В РЎвЂўР В РЎвЂ“Р В РЎвЂўР В Р вЂ Р РЋРЎвЂњР РЋР вЂ№ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РўвЂР В Р’В»Р РЋР РЏ Р В Р’В·Р В Р’В°Р В РЎвЂќР РЋР вЂљР РЋРІР‚в„–Р РЋРІР‚С™Р В РЎвЂР РЋР РЏ Р РЋР С“Р В РЎВР В Р’ВµР В Р вЂ¦Р РЋРІР‚в„–:"

        await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
        context.user_data['flow'] = 'closing'
        return CLOSE_SHIFT

    async def start_payment_image_upload(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ask cashier to choose Uzcard or Humo and then upload image"""
        if not await self._ensure_cashier_authenticated(update, context):
            return MAIN_MENU
        context.user_data.pop('blocked_media_group_id', None)

        lang = 'uz'
        user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await update.message.reply_text("Foydalanuvchi topilmadi.")
            return MAIN_MENU

        active_shift = self.db.fetch_one(
            "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE",
            (user_row['id'],)
        )
        if not active_shift:
            await update.message.reply_text("Ochiq smena yo'q.")
            return MAIN_MENU

        context.user_data['current_shift_id'] = active_shift['id']
        context.user_data['flow'] = 'payment_image'

        text = "Qaysi turdagi rasm yuborasiz?" if lang == 'uz' else "Qaysi turdagi rasm yuborasiz?"
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Uzcard", callback_data="payimg:uzcard")],
            [InlineKeyboardButton("Humo", callback_data="payimg:humo")],
            [InlineKeyboardButton("Orqaga", callback_data="payimg:back")],
        ])
        await update.message.reply_text(text, reply_markup=keyboard)
        return SELECT_PAYMENT_IMAGE

    async def select_payment_image_type(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        key = (query.data or '').split(':', 1)[1] if query.data else ''
        if key == 'back':
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            return MAIN_MENU
        if key not in ['uzcard', 'humo']:
            await self.start_payment_image_upload(update, context)
            return SELECT_PAYMENT_IMAGE

        context.user_data['pending_payment_image'] = key
        msg = "Rasmni yuboring:" if 'uz' == 'uz' else "Rasmni yuboring:"
        await context.bot.send_message(chat_id=query.message.chat_id, text=msg)
        return UPLOAD_PAYMENT_IMAGE

    async def upload_payment_image(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        try:
            if self._is_blocked_media_group(update, context):
                return UPLOAD_PAYMENT_IMAGE

            file_id = self._get_image_file_id(update)
            if not file_id:
                await update.message.reply_text("Iltimos, rasm yuboring (foto yoki image fayl).")
                return UPLOAD_PAYMENT_IMAGE

            key = context.user_data.get('pending_payment_image')
            shift_id = context.user_data.get('current_shift_id')
            if not shift_id:
                user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id=%s", (update.effective_user.id,))
                if user_row:
                    active_shift = self.db.fetch_one(
                        "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
                        (user_row['id'],)
                    )
                    if active_shift:
                        shift_id = active_shift['id']
                        context.user_data['current_shift_id'] = shift_id

            if key not in ('uzcard', 'humo') or not shift_id:
                await update.message.reply_text("Avval `Rasm jo'natish` tugmasini bosib, Uzcard yoki Humo ni tanlang.")
                context.user_data.pop('pending_payment_image', None)
                context.user_data['flow'] = None
                return MAIN_MENU

            image_type = 'uzcard_payment' if key == 'uzcard' else 'humo_payment'
            self.db.execute_query(
                "INSERT INTO images (shift_id, image_url, image_type) VALUES (%s, %s, %s)",
                (shift_id, file_id, image_type)
            )
            if key == 'uzcard':
                await self._send_group_shift_photo(
                    context, shift_id, file_id, "Uzcard hisobot rasmi", event_time=getattr(update.message, "date", None)
                )
            else:
                await self._send_group_shift_photo(
                    context, shift_id, file_id, "Humo hisobot rasmi", event_time=getattr(update.message, "date", None)
                )

            context.user_data.pop('pending_payment_image', None)
            context.user_data['flow'] = None
            self._block_current_media_group(update, context)
            if key == 'uzcard':
                await update.message.reply_text("Uzcard hisobot rasmingiz qabul qilindi.")
            else:
                await update.message.reply_text("Humo hisobot rasmingiz qabul qilindi.")

            uzcard_img = self._count_shift_images(shift_id, 'uzcard_payment')
            humo_img = self._count_shift_images(shift_id, 'humo_payment')
            if uzcard_img >= 1 and humo_img >= 1:
                await update.message.reply_text("Uzcard va Humo rasmlari to'liq qabul qilindi.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU
        except Exception:
            logger.exception("upload_payment_image failed")
            context.user_data.pop('pending_payment_image', None)
            context.user_data['flow'] = None
            await update.message.reply_text("Rasmni saqlashda xatolik bo'ldi. Qayta urinib ko'ring.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU
    async def edit_reports(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle report editing"""
        context.user_data['flow'] = 'edit'
        await self.show_edit_reports_menu(update, context)
        return EDIT_REPORT_SELECT

    async def show_edit_reports_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        lang = 'uz'
        user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Foydalanuvchi topilmadi.")
            context.user_data['flow'] = None
            return MAIN_MENU

        # Prefer current shift; fallback to latest shift
        shift_id = context.user_data.get('current_shift_id')
        if not shift_id:
            last_shift = self.db.fetch_one(
                "SELECT id FROM shifts WHERE user_id=%s ORDER BY id DESC LIMIT 1",
                (user_row['id'],)
            )
            shift_id = last_shift['id'] if last_shift else None

        report_row = None
        if shift_id:
            report_row = self.db.fetch_one(
                "SELECT * FROM reports WHERE shift_id=%s AND report_type='daily_report' ORDER BY id DESC LIMIT 1",
                (shift_id,)
            )
        context.user_data['edit_report_id'] = report_row['id'] if report_row else None

        # Build list of editable fields
        fields = []
        for key, label_uz, label_ru, *_rest in self._sverka_config():
            label = label_uz if lang == 'uz' else label_ru
            if report_row:
                value = report_row.get(key)
                fields.append((key, label, value))
            else:
                if key in context.user_data:
                    value = context.user_data.get(key)
                    fields.append((key, label, value))

        if not fields:
            msg = "Tahrirlash uchun hisobot topilmadi. Avval sverka tugating." if lang == 'uz' else "Tahrirlash uchun hisobot topilmadi. Avval sverka tugating."
            await context.bot.send_message(chat_id=update.effective_chat.id, text=msg)
            context.user_data['flow'] = None
            return MAIN_MENU

        text_lines = ["Tahrirlanadigan hisobotlar:" if lang == 'uz' else "Tahrirlanadigan hisobotlar:"]
        for _, label, value in fields:
            text_lines.append(f"- {label}: {value if value is not None else 0}")

        keyboard = []
        row = []
        for key, label, value in fields:
            btn = InlineKeyboardButton(f"{label}", callback_data=f"edit:{key}")
            row.append(btn)
            if len(row) == 2:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton("Orqaga", callback_data="edit:back")])

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="\n".join(text_lines),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return EDIT_REPORT_SELECT

    async def edit_reports_select(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        context.user_data['flow'] = 'edit'
        key = (query.data or '').split(':', 1)[1] if query.data else ''
        if key == 'back':
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            return MAIN_MENU

        config = {c[0]: c for c in self._sverka_config()}
        if key not in config:
            await self.show_edit_reports_menu(update, context)
            return EDIT_REPORT_SELECT

        _, label_uz, label_ru, *_rest = config[key]
        label = label_uz if 'uz' == 'uz' else label_ru
        context.user_data['pending_edit_key'] = key
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"{label} uchun yangi summani kiriting:"
        )
        return EDIT_REPORT_VALUE

    async def edit_reports_value(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        try:
            amount = self._parse_amount(update.message.text)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return EDIT_REPORT_VALUE

        key = context.user_data.get('pending_edit_key')
        if not key:
            await self.show_edit_reports_menu(update, context)
            return EDIT_REPORT_SELECT

        context.user_data.pop('pending_edit_key', None)
        context.user_data[key] = amount
        self._mark_sverka_done(context, key)

        report_id = context.user_data.get('edit_report_id')
        if report_id:
            self.db.execute_query(
                f"UPDATE reports SET {key}=%s WHERE id=%s",
                (amount, report_id)
            )

        await update.message.reply_text("Saqlab qo'yildi.")
        await self.show_edit_reports_menu(update, context)
        return EDIT_REPORT_SELECT

    async def send_reports(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ask for a date range and then show reports for that range."""
        context.user_data['admin_reports_range_pending'] = True
        await update.message.reply_text(
            "Qaysi vaqt oralig'ini ko'rasiz?\n"
            "Format:\n"
            "- 2026-03-01 2026-03-16\n"
            "yoki\n"
            "- 01.03.2026 16.03.2026"
        )
        return

    async def handle_reports_menu_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle report period selection from inline menu."""
        query = update.callback_query
        await query.answer()
        action = (query.data or "").split(":", 1)[-1]
        if action in ("daily", "weekly", "monthly"):
            await query.edit_message_text("Filialni tanlang:")
            await self._ask_report_location(update, context, action, chat_id=query.message.chat_id)
            return

        # custom date range
        context.user_data['admin_reports_range_pending'] = True
        await query.edit_message_text(
            "Qaysi vaqt oralig'ini ko'rasiz?\n"
            "Format:\n"
            "- 2026-03-01 2026-03-16\n"
            "yoki\n"
            "- 01.03.2026 16.03.2026"
        )

    async def _ask_report_location(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        period: str,
        chat_id: int = None,
    ):
        locations = self.db.fetch_all("SELECT id, name FROM locations WHERE is_active=TRUE ORDER BY id")
        keyboard = [[InlineKeyboardButton("Barcha filiallar", callback_data=f"reploc:all:{period}")]]
        for loc in locations:
            keyboard.append([InlineKeyboardButton(loc["name"], callback_data=f"reploc:{loc['id']}:{period}")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        target_chat = chat_id or update.effective_chat.id
        await context.bot.send_message(
            chat_id=target_chat,
            text="Qaysi filial hisobotini ko'rasiz?",
            reply_markup=reply_markup,
        )

    async def handle_report_location_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        parts = (query.data or "").split(":")
        if len(parts) != 3:
            await query.edit_message_text("Noto'g'ri tanlov.")
            return

        _, loc_raw, period = parts
        location_id = None if loc_raw == "all" else int(loc_raw)
        today = datetime.now().date()

        if period == "daily":
            start, end = today, today
        elif period == "weekly":
            start, end = today - timedelta(days=6), today
        elif period == "monthly":
            start, end = today - timedelta(days=29), today
        elif period == "range":
            rng = context.user_data.get("admin_reports_range_values")
            if not rng:
                await query.edit_message_text("Avval vaqt oralig'ini kiriting.")
                return
            start, end = rng
            context.user_data.pop("admin_reports_range_values", None)
        else:
            await query.edit_message_text("Noto'g'ri davr tanlandi.")
            return

        await query.edit_message_text("Hisobot tayyorlanmoqda...")
        await self._send_reports_for_range(
            query.message.chat_id,
            context,
            start,
            end,
            location_id=location_id,
        )

    async def _send_reports_for_range(self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, start, end, location_id=None):
        location_filter_sql = ""
        params = [str(start), str(end)]
        if location_id is not None:
            location_filter_sql = " AND s.location_id = %s"
            params.append(int(location_id))
        query = """
            SELECT
                r.id,
                u.first_name,
                u.last_name,
                l.name AS location,
                s.opened_at,
                s.closed_at,
                COALESCE(s.closing_amount,0) AS closing_amount,
                COALESCE(r.sales_amount,0) AS sales_amount,
                COALESCE(r.debt_received,0) AS debt_received,
                COALESCE(r.expenses,0) AS expenses,
                COALESCE(r.uzcard_amount,0) AS uzcard_amount,
                COALESCE(r.humo_amount,0) AS humo_amount,
                COALESCE(r.uzcard_refund,0) AS uzcard_refund,
                COALESCE(r.humo_refund,0) AS humo_refund,
                COALESCE(r.other_payments,0) AS other_payments,
                COALESCE(r.debt_payments,0) AS debt_payments,
                COALESCE(r.debt_refunds,0) AS debt_refunds
            FROM reports r
            JOIN shifts s ON r.shift_id = s.id
            JOIN users u ON s.user_id = u.id
            JOIN locations l ON s.location_id = l.id
            WHERE r.report_type='daily_report'
              AND DATE(s.opened_at) BETWEEN %s AND %s
        """ + location_filter_sql + """
            ORDER BY s.opened_at DESC
            LIMIT 200
        """
        rows = self.db.fetch_all(query, tuple(params))
        if not rows:
            # Daily so'rovda bugunda ma'lumot bo'lmasa oxirgi mavjud kunni ko'rsatamiz
            if str(start) == str(end):
                latest_q = """
                    SELECT DATE(s.opened_at) AS d
                    FROM reports r
                    JOIN shifts s ON r.shift_id = s.id
                    WHERE r.report_type='daily_report'
                """ + location_filter_sql + """
                    ORDER BY d DESC
                    LIMIT 1
                """
                latest = self.db.fetch_one(latest_q, tuple(params[2:]) if location_id is not None else ())
                if latest and latest.get('d'):
                    latest_day = str(latest['d'])
                    retry_params = [latest_day, latest_day]
                    if location_id is not None:
                        retry_params.append(int(location_id))
                    rows = self.db.fetch_all(query, tuple(retry_params))
                    if rows:
                        start = latest_day
                        end = latest_day
                        await context.bot.send_message(
                            chat_id=chat_id,
                            text=f"Bugungi hisobot topilmadi. Oxirgi mavjud sana ({latest_day}) ko'rsatildi."
                        )
            if not rows:
                await context.bot.send_message(chat_id=chat_id, text="Hisobotlar topilmadi.")
                return

        title = f"Hisobotlar ({start} - {end})"
        location_name = None
        if location_id is not None:
            loc = self.db.fetch_one("SELECT name FROM locations WHERE id=%s", (location_id,))
            if loc:
                location_name = loc['name']
                title += f" | Filial: {location_name}"
        lines = [title + ":"]

        def fmt(n):
            try:
                return f"{float(n or 0):,.0f}".replace(",", " ")
            except Exception:
                return str(n)

        for row in rows:
            cashier_name = f"{row['first_name']} {row['last_name'] or ''}".strip()
            total_balance = (
                row['sales_amount']
                + row['debt_received']
                + row['uzcard_amount']
                + row['humo_amount']
                + row['other_payments']
                + row['debt_refunds']
                - row['expenses']
                - row['debt_payments']
                - row['uzcard_refund']
                - row['humo_refund']
            )
            day = str(row['opened_at'])[:10]
            closed_at = str(row.get('closed_at') or '')[:16]
            closing_amount = fmt(row.get('closing_amount', 0))
            lines.append(
                f"{day} | {cashier_name} | {row['location']} | "
                f"Sof: {fmt(total_balance)} | Yopish: {closing_amount} | Yopilgan: {closed_at or '-'}"
            )

        msg = "\n".join(lines)
        if len(msg) > 3800:
            msg = msg[:3800] + "\n...(qisqartirildi)"
        await context.bot.send_message(chat_id=chat_id, text=msg)

        # Shu oralig' bo'yicha Excel ham yuboramiz
        try:
            xlsx = self._build_range_report_xlsx(rows, start, end, location_name)
            await context.bot.send_document(
                chat_id=chat_id,
                document=InputFile(xlsx, filename=f"hisobot_{start}_{end}.xlsx"),
                caption=title
            )
        except Exception:
            logger.exception("Range report Excel yuborishda xatolik")

    def _build_range_report_xlsx(self, rows, start, end, location_name=None) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hisobot"

        headers = [
            "Sana",
            "Kassir",
            "Filial",
            "Yopilgan vaqt",
            "Yopish summasi",
            "Savdo",
            "Kelgan qarz",
            "Chiqim",
            "Uzcard",
            "Humo",
            "Uzcard vozvrat",
            "Humo vozvrat",
            "Boshqa to'lovlar",
            "Qarzga berilgan to'lovlar",
            "Vozvrat qarzlar",
            "Sof summa",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _f(v):
            try:
                return float(v or 0)
            except Exception:
                return 0.0

        for row in rows:
            cashier_name = f"{row['first_name']} {row['last_name'] or ''}".strip()
            total_balance = (
                _f(row['sales_amount'])
                + _f(row['debt_received'])
                + _f(row['uzcard_amount'])
                + _f(row['humo_amount'])
                + _f(row['other_payments'])
                + _f(row['debt_refunds'])
                - _f(row['expenses'])
                - _f(row['debt_payments'])
                - _f(row['uzcard_refund'])
                - _f(row['humo_refund'])
            )
            ws.append([
                str(row['opened_at'])[:10],
                cashier_name,
                row['location'],
                str(row.get('closed_at') or '')[:19],
                _f(row.get('closing_amount')),
                _f(row['sales_amount']),
                _f(row['debt_received']),
                _f(row['expenses']),
                _f(row['uzcard_amount']),
                _f(row['humo_amount']),
                _f(row['uzcard_refund']),
                _f(row['humo_refund']),
                _f(row['other_payments']),
                _f(row['debt_payments']),
                _f(row['debt_refunds']),
                total_balance,
            ])

        for col in range(5, 17):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = "#,##0"

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 34)

        meta = wb.create_sheet("Ma'lumot")
        meta.append(["Boshlanish", str(start)])
        meta.append(["Tugash", str(end)])
        meta.append(["Filial", location_name or "Barcha filiallar"])
        meta.append(["Yaratilgan vaqt", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    def _parse_date_range(self, text: str):
        text = (text or "").strip()
        parts = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        fmt = "%Y-%m-%d"
        if len(parts) < 2:
            parts = re.findall(r"\d{2}[./]\d{2}[./]\d{4}", text)
            fmt = "%d.%m.%Y" if "." in (parts[0] if parts else "") else "%d/%m/%Y"
        if len(parts) < 2:
            return None
        try:
            start = datetime.strptime(parts[0].replace("/", "."), fmt).date()
            end = datetime.strptime(parts[1].replace("/", "."), fmt).date()
        except Exception:
            return None
        if start > end:
            start, end = end, start
        return start, end

    async def handle_admin_reports_range(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        rng = self._parse_date_range(update.message.text if update.message else "")
        if not rng:
            await update.message.reply_text(
                "Format noto'g'ri. Masalan:\n"
                "2026-03-01 2026-03-16\n"
                "yoki\n"
                "01.03.2026 16.03.2026"
            )
            return

        context.user_data['admin_reports_range_pending'] = False
        start, end = rng
        context.user_data['admin_reports_range_values'] = (start, end)
        await self._ask_report_location(update, context, "range")
    async def send_all_cashiers(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Send all cashiers to admin"""
        query = "SELECT * FROM users WHERE role = 'cashier' AND is_active = TRUE"
        cashiers = self.db.fetch_all(query)
        
        lang = 'uz'
        
        if lang == 'uz':
            if cashiers:
                msg = "Barcha kassirlar:\n"
                for cashier in cashiers:
                    msg += f"- {cashier['first_name']} {cashier['last_name']} ({cashier['phone_number']})\n"
            else:
                msg = "Hech qanday kassir topilmadi."
        else:
            if cashiers:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р В РЎвЂњР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“:\n"
                for cashier in cashiers:
                    msg += f"- {cashier['first_name']} {cashier['last_name']} ({cashier['phone_number']})\n"
            else:
                msg = "Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚В Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋРІР‚вЂњР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°."
                
        await update.message.reply_text(msg)

    async def handle_approval_requests(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle cashier approval requests"""
        query = "SELECT * FROM approval_requests WHERE status = 'pending'"
        requests = self.db.fetch_all(query)
        
        lang = 'uz'
        
        if lang == 'uz':
            if requests:
                msg = "Kassir so'rovlari:"
                for req in requests:
                    text = f"{req['first_name']} {req['last_name']} ({req['phone_number']}) | ID: {req['telegram_id']}"
                    keyboard = InlineKeyboardMarkup([
                        [
                            InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve:{req['telegram_id']}"),
                            InlineKeyboardButton("❌ Yo'q", callback_data=f"reject:{req['telegram_id']}")
                        ]
                    ])
                    await update.message.reply_text(text, reply_markup=keyboard)
            else:
                msg = "Yangi so'rovlar yo'q."
        else:
            if requests:
                msg = "Р В Р’В Р Р†Р вЂљРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В :"
                for req in requests:
                    text = f"{req['first_name']} {req['last_name']} ({req['phone_number']}) | ID: {req['telegram_id']}"
                    keyboard = InlineKeyboardMarkup([
                        [
                            InlineKeyboardButton("Р В Р вЂ Р РЋРЎв„ўР Р†Р вЂљР’В¦ Р В Р’В Р РЋРІР‚С”Р В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°", callback_data=f"approve:{req['telegram_id']}"),
                            InlineKeyboardButton("Р В Р вЂ Р РЋРЎС™Р В Р вЂ° Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В·", callback_data=f"reject:{req['telegram_id']}")
                        ]
                    ])
                    await update.message.reply_text(text, reply_markup=keyboard)
            else:
                msg = "Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р Р‹Р Р†Р вЂљР’В¦ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В ."
                
        if msg:
            await update.message.reply_text(msg)

    async def handle_approval_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle inline approve/reject callbacks"""
        query = update.callback_query
        await query.answer()

        data = query.data or ""
        if not (data.startswith("approve:") or data.startswith("reject:")):
            return

        # Only admins can approve/reject
        admin = self.db.fetch_one(
            "SELECT * FROM users WHERE telegram_id = %s AND role = 'admin' AND is_active = TRUE",
            (update.effective_user.id,)
        )
        if not admin:
            await query.edit_message_text("Faqat admin tasdiqlashi mumkin.")
            return

        try:
            telegram_id = int(data.split(":", 1)[1])
        except ValueError:
            await query.edit_message_text("Noto'g'ri ID.")
            return

        if data.startswith("approve:"):
            await self.approve_cashier(update, context, telegram_id)
            await query.edit_message_text("So'rov tasdiqlandi.")
        else:
            await self.reject_cashier(update, context, telegram_id)
            await query.edit_message_text("So'rov rad etildi.")

    async def notify_admins_new_request(self, context: ContextTypes.DEFAULT_TYPE, user_data: dict):
        """Notify all admins about new cashier approval request"""
        admins = self.db.fetch_all("SELECT telegram_id FROM users WHERE role = 'admin' AND is_active = TRUE")
        if not admins:
            return

        text = (
            "Yangi kassir so'rovi:\n"
            f"{user_data['first_name']} {user_data['last_name']} ({user_data['phone_number']})\n"
            f"ID: {user_data['telegram_id']}"
        )
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve:{user_data['telegram_id']}"),
                InlineKeyboardButton("❌ Yo'q", callback_data=f"reject:{user_data['telegram_id']}")
            ]
        ])

        for admin in admins:
            try:
                await context.bot.send_message(chat_id=admin['telegram_id'], text=text, reply_markup=keyboard)
            except Exception:
                continue

    def _extract_telegram_id(self, text: str):
        """Extract numeric telegram id from text."""
        match = re.search(r"\b(\d{5,20})\b", text)
        if match:
            try:
                return int(match.group(1))
            except ValueError:
                return None
        return None

    async def approve_cashier(self, update: Update, context: ContextTypes.DEFAULT_TYPE, telegram_id: int):
        """Approve cashier registration request"""
        # Check pending request
        query = "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'pending'"
        req = self.db.fetch_one(query, (telegram_id,))
        if not req:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Bunday pending so'rov topilmadi.")
            return

        # Insert or reactivate user
        user = self.db.fetch_one("SELECT * FROM users WHERE telegram_id = %s", (telegram_id,))
        if user:
            if not user.get('password_hash') and req.get('password_hash'):
                self.db.execute_query(
                    "UPDATE users SET role = 'cashier', is_active = TRUE, password_hash = %s WHERE telegram_id = %s",
                    (req.get('password_hash'), telegram_id)
                )
            else:
                self.db.execute_query(
                    "UPDATE users SET role = 'cashier', is_active = TRUE WHERE telegram_id = %s",
                    (telegram_id,)
                )
        else:
            self.db.execute_query(
                """
                INSERT INTO users (telegram_id, first_name, last_name, phone_number, role, password_hash, is_active)
                VALUES (%s, %s, %s, %s, 'cashier', %s, TRUE)
                """,
                (req['telegram_id'], req['first_name'], req['last_name'], req['phone_number'], req.get('password_hash'))
            )

        # Update request status
        self.db.execute_query(
            "UPDATE approval_requests SET status = 'approved', approved_at = NOW() WHERE telegram_id = %s",
            (telegram_id,)
        )

        # Notify cashier
        try:
            await context.bot.send_message(
                chat_id=telegram_id,
                text="Sizning so'rovingiz tasdiqlandi. Botdan foydalanishingiz mumkin."
            )
        except Exception:
            pass

        await context.bot.send_message(chat_id=update.effective_chat.id, text="So'rov tasdiqlandi.")

    async def reject_cashier(self, update: Update, context: ContextTypes.DEFAULT_TYPE, telegram_id: int):
        """Reject cashier registration request"""
        query = "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'pending'"
        req = self.db.fetch_one(query, (telegram_id,))
        if not req:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Bunday pending so'rov topilmadi.")
            return

        self.db.execute_query(
            "UPDATE approval_requests SET status = 'rejected', approved_at = NOW() WHERE telegram_id = %s",
            (telegram_id,)
        )

        try:
            await context.bot.send_message(
                chat_id=telegram_id,
                text="Sizning so'rovingiz rad etildi. Administrator bilan bog'laning."
            )
        except Exception:
            pass

        await context.bot.send_message(chat_id=update.effective_chat.id, text="So'rov rad etildi.")

    def _save_shift_image(self, shift_id: int, image_type: str, file_id: str):
        """Persist image reference for a shift"""
        self.db.execute_query(
            """
            INSERT INTO images (shift_id, image_url, image_type)
            VALUES (%s, %s, %s)
            """,
            (shift_id, file_id, image_type)
        )

    def _count_shift_images(self, shift_id: int, image_type: str) -> int:
        row = self.db.fetch_one(
            "SELECT COUNT(*) as cnt FROM images WHERE shift_id=%s AND image_type=%s",
            (shift_id, image_type)
        )
        return int(row['cnt']) if row else 0

    def _today_shift_for_user(self, user_id: int):
        """Foydalanuvchining bugungi (ochilgan sanasi bugun bo'lgan) oxirgi smenasi."""
        today = datetime.now().date().isoformat()
        return self.db.fetch_one(
            """
            SELECT id, is_open, opened_at, closed_at
            FROM shifts
            WHERE user_id=%s AND DATE(opened_at)=%s
            ORDER BY id DESC
            LIMIT 1
            """,
            (user_id, today)
        )

    def _get_image_file_id(self, update: Update):
        """Telegram'dan rasm file_id ni olish (photo yoki image document)."""
        msg = update.message
        if not msg:
            return None
        if msg.photo:
            return msg.photo[-1].file_id
        if msg.document:
            mime = (msg.document.mime_type or "").lower()
            name = (msg.document.file_name or "").lower()
            if mime.startswith("image/") or name.endswith((".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".heic")):
                return msg.document.file_id
        return None

    def _is_blocked_media_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
        """Agar oldingi bosqichda qabul qilingan albomning qolgan rasmlari kelsa, ularni e'tiborsiz qoldirish."""
        msg = update.message
        if not msg:
            return False
        media_group_id = getattr(msg, "media_group_id", None)
        blocked = context.user_data.get("blocked_media_group_id")
        return bool(media_group_id and blocked and str(media_group_id) == str(blocked))

    def _block_current_media_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Bosqich yakunlanganda shu albomning qolgan rasmlari keyingi bosqichga o'tib ketmasin."""
        msg = update.message
        if not msg:
            return
        media_group_id = getattr(msg, "media_group_id", None)
        if media_group_id:
            context.user_data["blocked_media_group_id"] = str(media_group_id)

    def _parse_amount(self, text: str) -> float:
        """Parse amounts like '12 300', '12,300', '12330 so'm'."""
        raw = text.strip()
        digits = ''.join(ch for ch in raw if ch.isdigit() or ch in ['.', ','])
        if not digits:
            return float(raw)
        return float(digits.replace(',', ''))

    def _format_telegram_time(self, dt_value) -> str:
        """Telegram message vaqtini Asia/Tashkent ga o'tkazib formatlaydi."""
        if not dt_value:
            return datetime.now(ZoneInfo("Asia/Tashkent")).strftime("%Y-%m-%d %H:%M:%S")
        try:
            return dt_value.astimezone(ZoneInfo("Asia/Tashkent")).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(dt_value)[:19]

    def _sverka_config(self):
        return [
            ('sales_amount', "Savdo summasi", "Р В Р Р‹Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В° Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР В РўвЂР В Р’В°Р В Р’В¶", REPORT_SALES, "Bugungi savdo miqdorini kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР В РўвЂР В Р’В°Р В Р’В¶:"),
            ('debt_received', "Kelgan qarzlar", "Р В РЎСџР РЋР вЂљР В РЎвЂР РЋРІвЂљВ¬Р В Р’ВµР В РўвЂР РЋРІвЂљВ¬Р В РЎвЂР В Р’Вµ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂ", REPORT_DEBT_RECEIVED, "Kelgan qarzlarni kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ”Р РЋР вЂљР В РЎвЂР РЋРІвЂљВ¬Р В Р’ВµР В РўвЂР РЋРІвЂљВ¬Р В РЎвЂР В Р’Вµ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂ (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('expenses', "Chiqimlar", "Р В Р’В Р В Р’В°Р РЋР С“Р РЋРІР‚В¦Р В РЎвЂўР В РўвЂР РЋРІР‚в„–", REPORT_EXPENSES, "Chiqimlarni kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР вЂљР В Р’В°Р РЋР С“Р РЋРІР‚В¦Р В РЎвЂўР В РўвЂР РЋРІР‚в„– (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('uzcard_amount', "Uzcard summasi", "Uzcard Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°", REPORT_UZCARD, "Uzcard orqali kiritilgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РЎвЂ”Р В РЎвЂў Uzcard (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('humo_amount', "Humo summasi", "Humo Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°", REPORT_HUMO, "Humo orqali kiritilgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РЎвЂ”Р В РЎвЂў Humo (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('uzcard_refund', "Uzcard vozvrat", "Р В РІР‚в„ўР В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Uzcard", REPORT_UZCARD_REFUND, "Uzcard orqali vozvrat bo'lgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РЎвЂ”Р В РЎвЂў Uzcard (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('humo_refund', "Humo vozvrat", "Р В РІР‚в„ўР В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Humo", REPORT_HUMO_REFUND, "Humo orqali vozvrat bo'lgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РЎвЂ”Р В РЎвЂў Humo (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('other_payments', "Boshqa to'lovlar", "Р В РІР‚СњР РЋР вЂљР РЋРЎвЂњР В РЎвЂ“Р В РЎвЂР В Р’Вµ Р В РЎвЂўР В РЎвЂ”Р В Р’В»Р В Р’В°Р РЋРІР‚С™Р РЋРІР‚в„–", REPORT_OTHER_PAYMENTS, "Boshqa to'lov turlarini kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РўвЂР РЋР вЂљР РЋРЎвЂњР В РЎвЂ“Р В РЎвЂР В Р’Вµ Р В РЎвЂўР В РЎвЂ”Р В Р’В»Р В Р’В°Р РЋРІР‚С™Р РЋРІР‚в„– (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('debt_payments', "Qarzga berilgan to'lovlar", "Р В РІР‚в„ўР РЋРІР‚в„–Р В РўвЂР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р вЂ  Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“", REPORT_DEBT_PAYMENTS, "Qarzga berilgan to'lovlarni kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р РЋРІР‚в„–Р В РўвЂР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р вЂ  Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“ (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('debt_refunds', "Vozvrat qarzlar", "Р В РІР‚в„ўР В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂўР В Р вЂ ", REPORT_DEBT_REFUNDS, "Vozvrat qarzlarni kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂўР В Р вЂ  (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):")
        ]

    def _opening_requirements_config(self):
        return [
            ("workplace_status", "Ish joyi holati rasmi", 2, UPLOAD_WORKPLACE_STATUS, "Ish stolingizni rasmga olib yuboring (2 ta rasm)."),
            ("terminal_power", "Terminal/ratsiya quvvati rasmi", 1, UPLOAD_TERMINAL_POWER, "Terminallar va ratsiyalar quvvat holatini rasmga oling."),
            ("zero_report", "Uzcard/Humo nol hisobot rasmi", 1, UPLOAD_ZERO_REPORT, "Uzcard va Humo nol hisobot rasmini yuboring."),
            ("opening_notification", "iiko/soliq ochilish rasmi", 1, UPLOAD_OPENING_NOTIFICATION, "iiko/soliq tizimida smena ochilganlik rasmini yuboring."),
            ("receipt_roll", "Zaxira chek lenta rasmi", 1, UPLOAD_RECEIPT_ROLL, "Zaxira chek lenta mavjudligi rasmini yuboring."),
        ]

    def _opening_missing_lines(self, shift_id: int):
        lines = []
        for key, label, required_count, *_ in self._opening_requirements_config():
            count = self._count_shift_images(shift_id, key)
            if count < required_count:
                remain = required_count - count
                if required_count == 1:
                    lines.append(f"- {label}")
                else:
                    lines.append(f"- {label} ({remain} ta qolgan)")
        return lines

    async def show_opening_requirements_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE, shift_id: int, note: str | None = None):
        cfg = self._opening_requirements_config()
        keyboard = []
        for key, label, required_count, *_ in cfg:
            count = self._count_shift_images(shift_id, key)
            done = count >= required_count
            icon = "✅" if done else "❌"
            suffix = ""
            if required_count > 1:
                suffix = f" ({count}/{required_count})"
            keyboard.append([InlineKeyboardButton(f"{icon} {label}{suffix}", callback_data=f"op:{key}")])

        keyboard.append([
            InlineKeyboardButton("Yangilash", callback_data="op:refresh"),
            InlineKeyboardButton("Orqaga", callback_data="op:back"),
        ])

        missing = self._opening_missing_lines(shift_id)
        text = note or "Smena ochish rasmlari holati:"
        if missing:
            text += "\n\nYetishmayotganlar:\n" + "\n".join(missing)
        else:
            text += "\n\nBarcha kerakli rasmlar yuklangan."

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard),
        )

    async def opening_select_step(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        if not query or not query.data:
            return MAIN_MENU
        await query.answer()

        key = query.data.split(":", 1)[1] if ":" in query.data else ""

        user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id=%s", (update.effective_user.id,))
        if not user_row:
            await context.bot.send_message(chat_id=query.message.chat_id, text="Foydalanuvchi topilmadi.")
            return MAIN_MENU
        active_shift = self.db.fetch_one(
            "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
            (user_row['id'],)
        )
        if not active_shift:
            await context.bot.send_message(chat_id=query.message.chat_id, text="Ochiq smena topilmadi.")
            return MAIN_MENU

        shift_id = int(active_shift["id"])
        context.user_data["current_shift_id"] = shift_id
        context.user_data["flow"] = "opening"

        if key == "back":
            await self.show_cashier_menu(update, context)
            return MAIN_MENU
        if key == "refresh":
            await self.show_opening_requirements_menu(update, context, shift_id)
            return MAIN_MENU

        cfg = {c[0]: c for c in self._opening_requirements_config()}
        if key not in cfg:
            await self.show_opening_requirements_menu(update, context, shift_id)
            return MAIN_MENU

        _, label, required_count, state, prompt = cfg[key]
        current_count = self._count_shift_images(shift_id, key)
        if current_count >= required_count:
            await context.bot.send_message(chat_id=query.message.chat_id, text=f"✅ {label} allaqachon bajarilgan.")
            await self.show_opening_requirements_menu(update, context, shift_id)
            return MAIN_MENU

        context.user_data["opening_stage"] = key
        await context.bot.send_message(chat_id=query.message.chat_id, text=prompt, reply_markup=ReplyKeyboardRemove())
        return state

    def _init_sverka_status(self, context: ContextTypes.DEFAULT_TYPE):
        status = context.user_data.get('sverka_status')
        if not isinstance(status, dict):
            status = {}
        for key, *_ in self._sverka_config():
            if key not in status:
                status[key] = bool(context.user_data.get(key) is not None)
            elif not status.get(key) and context.user_data.get(key) is not None:
                status[key] = True
        context.user_data['sverka_status'] = status

    def _mark_sverka_done(self, context: ContextTypes.DEFAULT_TYPE, key: str):
        self._init_sverka_status(context)
        context.user_data['sverka_status'][key] = True

    def _sverka_all_done(self, context: ContextTypes.DEFAULT_TYPE) -> bool:
        self._init_sverka_status(context)
        return all(context.user_data['sverka_status'].get(k, False) for k, *_ in self._sverka_config())

    def _invalid_amount_msg(self, context: ContextTypes.DEFAULT_TYPE) -> str:
        lang = 'uz'
        if lang == 'uz':
            return "Iltimos, to'g'ri miqdor kiriting."
        return "Р В РЎСџР В РЎвЂўР В Р’В¶Р В Р’В°Р В Р’В»Р РЋРЎвЂњР В РІвЂћвЂ“Р РЋР С“Р РЋРІР‚С™Р В Р’В°, Р В Р вЂ Р В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В РЎвЂР В Р’В»Р РЋР Р‰Р В Р вЂ¦Р РЋРЎвЂњР РЋР вЂ№ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ."

    async def show_sverka_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE, note: str | None = None):
        lang = 'uz'
        self._init_sverka_status(context)
        status = context.user_data.get('sverka_status', {})

        buttons = []
        for key, label_uz, label_ru, *_rest in self._sverka_config():
            label = label_uz if lang == 'uz' else label_ru
            icon = "✅" if status.get(key) else "❌"
            buttons.append(InlineKeyboardButton(f"{icon} {label}", callback_data=f"sv:{key}"))

        keyboard = []
        for i in range(0, len(buttons), 2):
            keyboard.append(buttons[i:i+2])

        finish_text = "Yakunlash" if lang == 'uz' else "Yakunlash"
        keyboard.append([InlineKeyboardButton(finish_text, callback_data="sv:finish")])

        if not note:
            note = "Sverka bo'limlarini tanlang:" if lang == 'uz' else "Sverka bo'limlarini tanlang:"

        # Add missing items list at the bottom
        missing = []
        for key, label_uz, label_ru, *_rest in self._sverka_config():
            if not status.get(key):
                missing.append(label_uz if lang == 'uz' else label_ru)
        if missing:
            note += "\n\nTo'ldirilmagan bandlar:\n- " + "\n- ".join(missing)

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=note,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    async def sverka_select_step(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()

        key = (query.data or '').split(':', 1)[1] if query.data else ''
        if key == 'finish':
            if not self._sverka_all_done(context):
                msg = "Hamma band to'ldirilmagan. Iltimos, qolganlarini to'ldiring." if 'uz' == 'uz' else "Р В РЎСљР В Р’Вµ Р В Р вЂ Р РЋР С“Р В Р’Вµ Р В РЎвЂ”Р РЋРЎвЂњР В Р вЂ¦Р В РЎвЂќР РЋРІР‚С™Р РЋРІР‚в„– Р В Р’В·Р В Р’В°Р В РЎвЂ”Р В РЎвЂўР В Р’В»Р В Р вЂ¦Р В Р’ВµР В Р вЂ¦Р РЋРІР‚в„–. Р В РІР‚вЂќР В Р’В°Р В РЎвЂ”Р В РЎвЂўР В Р’В»Р В Р вЂ¦Р В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂўР РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ Р РЋРІвЂљВ¬Р В РЎвЂР В Р’ВµР РЋР С“Р РЋР РЏ."
                await context.bot.send_message(chat_id=query.message.chat_id, text=msg)
                await self.show_sverka_menu(update, context)
                return SUBMIT_DAILY_REPORT
            await self.save_daily_report(update, context)
            msg = "Sverka yakunlandi! Barcha hisobotlar saqlandi." if 'uz' == 'uz' else "Р В Р Р‹Р В Р вЂ Р В Р’ВµР РЋР вЂљР В РЎвЂќР В Р’В° Р В Р’В·Р В Р’В°Р В Р вЂ Р В Р’ВµР РЋР вЂљР РЋРІвЂљВ¬Р В Р’ВµР В Р вЂ¦Р В Р’В°! Р В РІР‚в„ўР РЋР С“Р В Р’Вµ Р В РЎвЂўР РЋРІР‚С™Р РЋРІР‚РЋР В Р’ВµР РЋРІР‚С™Р РЋРІР‚в„– Р РЋР С“Р В РЎвЂўР РЋРІР‚В¦Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В Р’ВµР В Р вЂ¦Р РЋРІР‚в„–."
            await context.bot.send_message(chat_id=query.message.chat_id, text=msg)
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            return MAIN_MENU

        config = {c[0]: c for c in self._sverka_config()}
        if key not in config:
            await self.show_sverka_menu(update, context)
            return SUBMIT_DAILY_REPORT

        _, label_uz, label_ru, state, prompt_uz, prompt_ru = config[key]
        context.user_data['pending_sverka_key'] = key
        context.user_data['pending_sverka_state'] = state
        prompt = prompt_uz if 'uz' == 'uz' else prompt_ru
        await context.bot.send_message(chat_id=query.message.chat_id, text=prompt, reply_markup=ReplyKeyboardRemove())
        return state

    async def _after_sverka_step(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if self._sverka_all_done(context):
            await self.save_daily_report(update, context)
            msg = "Sverka yakunlandi! Barcha hisobotlar saqlandi." if 'uz' == 'uz' else "Р В Р Р‹Р В Р вЂ Р В Р’ВµР РЋР вЂљР В РЎвЂќР В Р’В° Р В Р’В·Р В Р’В°Р В Р вЂ Р В Р’ВµР РЋР вЂљР РЋРІвЂљВ¬Р В Р’ВµР В Р вЂ¦Р В Р’В°! Р В РІР‚в„ўР РЋР С“Р В Р’Вµ Р В РЎвЂўР РЋРІР‚С™Р РЋРІР‚РЋР В Р’ВµР РЋРІР‚С™Р РЋРІР‚в„– Р РЋР С“Р В РЎвЂўР РЋРІР‚В¦Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В Р’ВµР В Р вЂ¦Р РЋРІР‚в„–."
            await update.message.reply_text(msg)
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            return MAIN_MENU

        note = "Qabul qilindi. Keyingi bandni tanlang." if 'uz' == 'uz' else "Р В РЎСџР РЋР вЂљР В РЎвЂР В Р вЂ¦Р РЋР РЏР РЋРІР‚С™Р В РЎвЂў. Р В РІР‚в„ўР РЋРІР‚в„–Р В Р’В±Р В Р’ВµР РЋР вЂљР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р В Р’В»Р В Р’ВµР В РўвЂР РЋРЎвЂњР РЋР вЂ№Р РЋРІР‚В°Р В РЎвЂР В РІвЂћвЂ“ Р В РЎвЂ”Р РЋРЎвЂњР В Р вЂ¦Р В РЎвЂќР РЋРІР‚С™."
        await self.show_sverka_menu(update, context, note=note)
        return SUBMIT_DAILY_REPORT

    async def close_shift(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Close active shift"""
        try:
            try:
                amount = self._parse_amount(update.message.text)
            except ValueError:
                await update.message.reply_text("Iltimos, to'g'ri miqdor kiriting.")
                return CLOSE_SHIFT

            user_row = self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
            if not user_row:
                await update.message.reply_text("Foydalanuvchi topilmadi.")
                return MAIN_MENU

            open_shifts = self.db.fetch_all(
                "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC",
                (user_row['id'],)
            ) or []
            if not open_shifts:
                await update.message.reply_text("Ochiq smena topilmadi.")
                return MAIN_MENU

            # Asosiy yopiladigan smena: eng oxirgisi
            shift_id = int(open_shifts[0]['id'])
            context.user_data['current_shift_id'] = shift_id

            self.db.execute_query(
                "UPDATE shifts SET closing_amount=%s, closed_at=NOW(), is_open=FALSE WHERE id=%s",
                (amount, shift_id)
            )

            # Xavfsizlik: agar tasodifan bir nechta ochiq smena qolgan bo'lsa, ularni ham yopamiz
            stale_ids = [int(r['id']) for r in open_shifts[1:]]
            if stale_ids:
                for sid in stale_ids:
                    self.db.execute_query(
                        "UPDATE shifts SET closed_at=COALESCE(closed_at, NOW()), is_open=FALSE WHERE id=%s",
                        (sid,)
                    )

            # Send full shift report + images report to group
            try:
                shift_ctx = self.db.fetch_one(
                    """
                    SELECT s.opened_at, s.closed_at, l.name AS location, u.first_name, u.last_name
                    FROM shifts s
                    JOIN users u ON s.user_id = u.id
                    JOIN locations l ON s.location_id = l.id
                    WHERE s.id=%s
                    """,
                    (shift_id,),
                ) or {}
                cashier_name = f"{shift_ctx.get('first_name','')} {shift_ctx.get('last_name') or ''}".strip() or update.effective_user.first_name
                location = shift_ctx.get("location") or ""
                report_date = str(shift_ctx.get("opened_at") or "")[:10] or datetime.now().strftime("%Y-%m-%d")
                closed_at = str(shift_ctx.get("closed_at") or "")[:19]

                full_xlsx = self._build_shift_full_xlsx(shift_id)
                await self._send_group_document(
                    context,
                    full_xlsx,
                    f"kunlik_kassir_hisobot_shift_{shift_id}.xlsx",
                    caption=(
                        "Kunlik kassir hisobot (Excel)\n"
                        f"Kassir: {cashier_name}\n"
                        f"Filial: {location}\n"
                        f"Sana: {report_date}"
                    )
                )

                images_xlsx = await self._build_shift_images_xlsx(context, shift_id)
                await self._send_group_document(
                    context,
                    images_xlsx,
                    f"rasmlar_shift_{shift_id}.xlsx",
                    caption=(
                        "Smena rasmlari (Excel)\n"
                        f"Kassir: {cashier_name}\n"
                        f"Filial: {location}\n"
                        f"Sana: {report_date}"
                    )
                )

                opening_images_xlsx = await self._build_opening_images_xlsx(context, shift_id)
                await self._send_group_document(
                    context,
                    opening_images_xlsx,
                    f"smena_ochish_rasmlari_shift_{shift_id}.xlsx",
                    caption=(
                        "Smena ochish rasmlari (Excel)\n"
                        f"Kassir: {cashier_name}\n"
                        f"Filial: {location}\n"
                        f"Sana: {report_date}"
                    )
                )

                await self._send_group_message(
                    context,
                    (
                        "Smena yopildi.\n"
                        f"Kassir: {cashier_name}\n"
                        f"Filial: {location}\n"
                        f"Yopilgan vaqt: {closed_at or '-'}\n"
                        f"Yopish summasi: {amount:,.0f}"
                    ).replace(",", " ")
                )
            except Exception:
                logger.exception("close_shift group send failed")

            await update.message.reply_text("Smena yopildi.")
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            context.user_data.pop('current_shift_id', None)
            context.user_data.pop('opening_stage', None)
            context.user_data.pop('pending_payment_image', None)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            return MAIN_MENU
        except Exception:
            logger.exception("close_shift failed")
            context.user_data['flow'] = None
            await update.message.reply_text("Smena yopishda xatolik bo'ldi. Qayta urinib ko'ring.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU

    def _build_shift_full_xlsx(self, shift_id: int) -> BytesIO:
        """
        Build one Excel file with all cashier data for the shift:
        - Smena (opened/closed, opening/closing amount, cashier, location)
        - Sverka (all numeric fields, one row)
        - Rasmlar (required photos + payment photos counts and file_ids)
        """
        shift = self.db.fetch_one(
            """
            SELECT
              s.id, s.opened_at, s.closed_at, s.opening_amount, s.closing_amount, s.is_open,
              u.first_name, u.last_name, u.phone_number,
              l.name AS location
            FROM shifts s
            JOIN users u ON s.user_id=u.id
            JOIN locations l ON s.location_id=l.id
            WHERE s.id=%s
            """,
            (shift_id,)
        ) or {}

        report = self.db.fetch_one(
            "SELECT * FROM reports WHERE shift_id=%s AND report_type='daily_report' ORDER BY id DESC LIMIT 1",
            (shift_id,)
        ) or {}

        images = self.db.fetch_all(
            """
            SELECT image_type, image_url, uploaded_at
            FROM images
            WHERE shift_id=%s
            ORDER BY uploaded_at ASC
            """,
            (shift_id,)
        ) or []

        wb = Workbook()
        ws_shift = wb.active
        ws_shift.title = "Smena"

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)

        def _set_kv(row_idx: int, key: str, value):
            ws_shift.cell(row=row_idx, column=1, value=key).font = bold
            ws_shift.cell(row=row_idx, column=2, value=value)

        cashier_name = (f"{shift.get('first_name','')} {shift.get('last_name') or ''}".strip() or "Kassir")
        _set_kv(1, "Kassir", cashier_name)
        _set_kv(2, "Telefon", shift.get("phone_number") or "")
        _set_kv(3, "Filial", shift.get("location") or "")
        _set_kv(4, "Smena ochilgan vaqt", str(shift.get("opened_at") or ""))
        _set_kv(5, "Smena yopilgan vaqt", str(shift.get("closed_at") or ""))
        _set_kv(6, "Ochilish summasi", float(shift.get("opening_amount") or 0))
        _set_kv(7, "Yopish summasi", float(shift.get("closing_amount") or 0))

        ws_shift.column_dimensions["A"].width = 22
        ws_shift.column_dimensions["B"].width = 42
        for r in range(1, 8):
            ws_shift.cell(row=r, column=1).alignment = Alignment(vertical="center")
            ws_shift.cell(row=r, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws_shift.cell(row=6, column=2).number_format = "#,##0"
        ws_shift.cell(row=7, column=2).number_format = "#,##0"

        # Sverka sheet
        ws_rep = wb.create_sheet("Sverka")
        rep_headers = [
            "Savdo", "Kelgan qarz", "Chiqim", "Uzcard", "Humo",
            "Uzcard vozvrat", "Humo vozvrat", "Boshqa to'lovlar",
            "Qarzga berilgan to'lovlar", "Vozvrat qarzlar", "Sof summa",
        ]
        ws_rep.append(rep_headers)

        def _f(key: str) -> float:
            try:
                return float(report.get(key) or 0)
            except Exception:
                return 0.0

        total_balance = (
            _f("sales_amount")
            + _f("debt_received")
            + _f("uzcard_amount")
            + _f("humo_amount")
            + _f("other_payments")
            + _f("debt_refunds")
            - _f("expenses")
            - _f("debt_payments")
            - _f("uzcard_refund")
            - _f("humo_refund")
        )

        ws_rep.append([
            _f("sales_amount"),
            _f("debt_received"),
            _f("expenses"),
            _f("uzcard_amount"),
            _f("humo_amount"),
            _f("uzcard_refund"),
            _f("humo_refund"),
            _f("other_payments"),
            _f("debt_payments"),
            _f("debt_refunds"),
            float(total_balance),
        ])

        for c in range(1, len(rep_headers) + 1):
            cell = ws_rep.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_rep.column_dimensions[get_column_letter(c)].width = min(max(12, len(rep_headers[c - 1]) + 2), 28)
            ws_rep.cell(row=2, column=c).number_format = "#,##0"

        # Images sheet
        ws_img = wb.create_sheet("Rasmlar")
        img_headers = ["Rasm turi", "Nechta", "Oxirgi vaqt", "File ID(lar)"]
        ws_img.append(img_headers)
        for c in range(1, len(img_headers) + 1):
            cell = ws_img.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_img.column_dimensions["A"].width = 22
        ws_img.column_dimensions["B"].width = 10
        ws_img.column_dimensions["C"].width = 20
        ws_img.column_dimensions["D"].width = 70

        # Group images by type
        by_type = {}
        for row in images:
            t = row.get("image_type")
            by_type.setdefault(t, []).append(row)

        type_labels = {
            "workplace_status": "Ish joyi holati (2 ta)",
            "terminal_power": "Terminal/ratsiya quvvati",
            "zero_report": "Nol hisobot (Uzcard/Humo)",
            "opening_notification": "Iiko/soliq smena ochildi",
            "receipt_roll": "Zaxira chek lenta",
            "uzcard_payment": "Uzcard to'lov rasmi",
            "humo_payment": "Humo to'lov rasmi",
        }
        ordered_types = [
            "workplace_status",
            "terminal_power",
            "zero_report",
            "opening_notification",
            "receipt_roll",
            "uzcard_payment",
            "humo_payment",
        ]

        r = 2
        for t in ordered_types:
            rows = by_type.get(t, [])
            cnt = len(rows)
            last_time = str(rows[-1].get("uploaded_at")) if rows else ""
            file_ids = ", ".join([str(x.get("image_url") or "") for x in rows if x.get("image_url")])
            ws_img.append([type_labels.get(t, t), cnt, last_time, file_ids])
            ws_img.cell(row=r, column=2).alignment = Alignment(horizontal="center")
            ws_img.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
            r += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    async def _build_opening_images_xlsx(self, context: ContextTypes.DEFAULT_TYPE, shift_id: int) -> BytesIO:
        """
        Build one Excel with only shift-opening images:
        - Ish joyi holati
        - Terminal/ratsiya holati
        - Nol hisobot
        - iiko/soliq ochilish
        - Zaxira chek lenta
        """
        opening_types = (
            "workplace_status",
            "terminal_power",
            "zero_report",
            "opening_notification",
            "receipt_roll",
        )
        rows = self.db.fetch_all(
            """
            SELECT image_type, image_url, uploaded_at
            FROM images
            WHERE shift_id=%s AND image_type IN (%s, %s, %s, %s, %s)
            ORDER BY uploaded_at ASC
            """,
            (shift_id, *opening_types),
        ) or []

        wb = Workbook()
        ws = wb.active
        ws.title = "Smena ochish rasmlari"

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")

        for col, label in enumerate(["Rasm nomi", "Rasm", "Sana/Vaqt"], start=1):
            c = ws.cell(row=1, column=col, value=label)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 24

        readable = {
            "workplace_status": "Ish joyi holati",
            "terminal_power": "Terminal/ratsiya holati",
            "zero_report": "Nol hisobot (Uzcard/Humo)",
            "opening_notification": "iiko/soliq ochilish bildirishnomasi",
            "receipt_roll": "Zaxira chek lenta",
        }

        async def add_img(ws_obj, cell, file_id):
            try:
                tg_file = await context.bot.get_file(file_id)
                data = await tg_file.download_as_bytearray()
                bio = BytesIO(data)
                img = PILImage.open(bio)
                out = BytesIO()
                img.convert("RGB").save(out, format="JPEG", quality=85)
                out.seek(0)
                xl_img = XLImage(out)
                xl_img.width = 260
                xl_img.height = 180
                ws_obj.add_image(xl_img, cell)
                return True
            except Exception:
                return False

        if not rows:
            ws.cell(row=2, column=1, value="Smena ochish rasmlari topilmadi.")
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        r_idx = 2
        for item in rows:
            ws.row_dimensions[r_idx].height = 140
            t = (item.get("image_type") or "").strip()
            ws.cell(row=r_idx, column=1, value=readable.get(t, t))
            ok = await add_img(ws, f"B{r_idx}", item["image_url"])
            if not ok:
                ws.cell(row=r_idx, column=2, value=f"Yuklab bo'lmadi: {item['image_url']}")
            ws.cell(row=r_idx, column=3, value=str(item.get("uploaded_at") or "")[:19])
            r_idx += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    async def _build_shift_images_xlsx(self, context: ContextTypes.DEFAULT_TYPE, shift_id: int) -> BytesIO:
        """
        Build one Excel with real images:
        - To'lov rasmlari: Uzcard | Humo (yonma-yon)
        - Ish jarayoni rasmlari: qolgan rasmlar ketma-ket
        """
        rows = self.db.fetch_all(
            """
            SELECT image_type, image_url, uploaded_at
            FROM images
            WHERE shift_id=%s
            ORDER BY uploaded_at ASC
            """,
            (shift_id,),
        ) or []

        wb = Workbook()
        ws_pay = wb.active
        ws_pay.title = "To'lov rasmlari"

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")

        for col, label in enumerate(["Uzcard rasm", "Humo rasm", "Sana/Vaqt"], start=1):
            c = ws_pay.cell(row=1, column=col, value=label)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws_pay.column_dimensions["A"].width = 42
        ws_pay.column_dimensions["B"].width = 42
        ws_pay.column_dimensions["C"].width = 24

        pay = defaultdict(list)
        other = []
        for r in rows:
            t = (r.get("image_type") or "").strip()
            if t in ("uzcard_payment", "humo_payment"):
                pay[t].append(r)
            else:
                other.append(r)

        async def add_img(ws, cell, file_id):
            try:
                tg_file = await context.bot.get_file(file_id)
                data = await tg_file.download_as_bytearray()
                bio = BytesIO(data)
                # Openpyxl barqaror ishlashi uchun PIL orqali qayta saqlaymiz
                img = PILImage.open(bio)
                out = BytesIO()
                img.convert("RGB").save(out, format="JPEG", quality=85)
                out.seek(0)
                xl_img = XLImage(out)
                xl_img.width = 260
                xl_img.height = 180
                ws.add_image(xl_img, cell)
                return True
            except Exception:
                return False

        uz_rows = pay.get("uzcard_payment", [])
        hu_rows = pay.get("humo_payment", [])
        max_len = max(len(uz_rows), len(hu_rows), 1)

        for i in range(max_len):
            excel_row = i + 2
            ws_pay.row_dimensions[excel_row].height = 140

            uz = uz_rows[i] if i < len(uz_rows) else None
            hu = hu_rows[i] if i < len(hu_rows) else None

            if uz:
                ok = await add_img(ws_pay, f"A{excel_row}", uz["image_url"])
                if not ok:
                    ws_pay.cell(row=excel_row, column=1, value=f"Yuklab bo'lmadi: {uz['image_url']}")
            if hu:
                ok = await add_img(ws_pay, f"B{excel_row}", hu["image_url"])
                if not ok:
                    ws_pay.cell(row=excel_row, column=2, value=f"Yuklab bo'lmadi: {hu['image_url']}")

            stamp = (uz or hu or {}).get("uploaded_at")
            ws_pay.cell(row=excel_row, column=3, value=str(stamp)[:19] if stamp else "")

        ws_other = wb.create_sheet("Ish jarayoni rasmlari")
        for col, label in enumerate(["Rasm turi", "Rasm", "Sana/Vaqt"], start=1):
            c = ws_other.cell(row=1, column=col, value=label)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws_other.column_dimensions["A"].width = 32
        ws_other.column_dimensions["B"].width = 42
        ws_other.column_dimensions["C"].width = 24

        readable = {
            "workplace_status": "Ish joyi holati",
            "terminal_power": "Terminal/ratsiya holati",
            "zero_report": "Nol hisobot",
            "opening_notification": "iiko/soliq ochilish",
            "receipt_roll": "Zaxira chek lenta",
        }

        r_idx = 2
        for item in other:
            ws_other.row_dimensions[r_idx].height = 140
            t = item.get("image_type") or ""
            ws_other.cell(row=r_idx, column=1, value=readable.get(t, t))
            ok = await add_img(ws_other, f"B{r_idx}", item["image_url"])
            if not ok:
                ws_other.cell(row=r_idx, column=2, value=f"Yuklab bo'lmadi: {item['image_url']}")
            ws_other.cell(row=r_idx, column=3, value=str(item.get("uploaded_at") or "")[:19])
            r_idx += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out
    async def reset_cashier_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE, telegram_id: int):
        """Reset cashier password and ask them to set a new one"""
        user = self.db.fetch_one(
            "SELECT * FROM users WHERE telegram_id = %s AND role = 'cashier'",
            (telegram_id,)
        )
        if not user:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Kassir topilmadi.")
            return

        self.db.execute_query(
            "UPDATE users SET password_hash = NULL WHERE telegram_id = %s",
            (telegram_id,)
        )

        try:
            await context.bot.send_message(
                chat_id=telegram_id,
                text="Parolingiz reset qilindi. /start bosing va yangi parol kiriting."
            )
        except Exception:
            pass

        await context.bot.send_message(chat_id=update.effective_chat.id, text="Parol reset qilindi.")

    async def _send_group_message(self, context: ContextTypes.DEFAULT_TYPE, text: str):
        if not GROUP_CHAT_ID:
            return
        try:
            await context.bot.send_message(chat_id=GROUP_CHAT_ID, text=text)
        except Exception:
            pass

    async def _send_group_photo(self, context: ContextTypes.DEFAULT_TYPE, file_id: str, caption: str = ""):
        if not GROUP_CHAT_ID:
            return
        try:
            await context.bot.send_photo(chat_id=GROUP_CHAT_ID, photo=file_id, caption=caption)
        except Exception:
            pass

    async def _send_group_document(self, context: ContextTypes.DEFAULT_TYPE, data: BytesIO, filename: str, caption: str = ""):
        if not GROUP_CHAT_ID:
            return
        try:
            data.seek(0)
            doc = InputFile(data, filename=filename)
            await context.bot.send_document(chat_id=GROUP_CHAT_ID, document=doc, caption=caption)
        except Exception:
            pass

    def _get_shift_meta(self, shift_id: int):
        row = self.db.fetch_one(
            """
            SELECT
                s.opened_at,
                l.name AS location,
                u.first_name,
                u.last_name
            FROM shifts s
            JOIN users u ON s.user_id = u.id
            JOIN locations l ON s.location_id = l.id
            WHERE s.id=%s
            """,
            (shift_id,),
        )
        if not row:
            return {"cashier": "", "location": "", "opened_at": ""}
        cashier = f"{row.get('first_name','')} {row.get('last_name') or ''}".strip()
        return {
            "cashier": cashier,
            "location": row.get("location") or "",
            "opened_at": str(row.get("opened_at") or "")[:19],
        }

    async def _send_group_shift_photo(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        shift_id: int,
        file_id: str,
        image_title: str,
        event_time=None
    ):
        # Talab bo'yicha guruhga alohida rasmlar yuborilmaydi.
        # Rasmlar faqat bitta/yig'ma fayl ko'rinishida smena yopilganda yuboriladi.
        return

    def _fmt_money(self, value) -> str:
        try:
            num = float(value or 0)
        except Exception:
            return str(value)
        return f"{num:,.0f}".replace(",", " ")

    def _build_sverka_xlsx(self, cashier_name: str, phone: str, location: str, opened_at, report_data: dict) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sverka"

        headers = [
            "Kassir", "Telefon", "Filial", "Smena ochilgan vaqt",
            "Savdo", "Kelgan qarz", "Chiqim", "Uzcard", "Humo",
            "Uzcard vozvrat", "Humo vozvrat", "Boshqa to'lovlar",
            "Qarzga berilgan to'lovlar", "Vozvrat qarzlar", "Sof summa",
        ]

        total_balance = (
            float(report_data.get("sales_amount", 0) or 0)
            + float(report_data.get("debt_received", 0) or 0)
            + float(report_data.get("uzcard_amount", 0) or 0)
            + float(report_data.get("humo_amount", 0) or 0)
            + float(report_data.get("other_payments", 0) or 0)
            + float(report_data.get("debt_refunds", 0) or 0)
            - float(report_data.get("expenses", 0) or 0)
            - float(report_data.get("debt_payments", 0) or 0)
            - float(report_data.get("uzcard_refund", 0) or 0)
            - float(report_data.get("humo_refund", 0) or 0)
        )

        row = [
            cashier_name,
            phone,
            location,
            str(opened_at),
            float(report_data.get("sales_amount", 0) or 0),
            float(report_data.get("debt_received", 0) or 0),
            float(report_data.get("expenses", 0) or 0),
            float(report_data.get("uzcard_amount", 0) or 0),
            float(report_data.get("humo_amount", 0) or 0),
            float(report_data.get("uzcard_refund", 0) or 0),
            float(report_data.get("humo_refund", 0) or 0),
            float(report_data.get("other_payments", 0) or 0),
            float(report_data.get("debt_payments", 0) or 0),
            float(report_data.get("debt_refunds", 0) or 0),
            float(total_balance),
        ]

        ws.append(headers)
        ws.append(row)

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        money_cols = list(range(5, 16))
        for col in money_cols:
            ws.cell(row=2, column=col).number_format = "#,##0"

        # Auto width
        for col in range(1, len(headers) + 1):
            max_len = 0
            for r in (1, 2):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 40)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out
    async def modify_user_data(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle user data modification"""
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Foydalanuvchi ma'lumotlarini o'zgartirish funksiyasi ishga tushirildi."
        else:
            msg = "Р В Р’В Р вЂ™Р’В¤Р В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљР’В Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В Р РЏ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В·Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В Р РЏ Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’В°Р В Р’В Р В РІР‚В¦Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р Р‹Р Р†Р вЂљР’В¦ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°Р В Р’В Р вЂ™Р’В·Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’ВµР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р РЏ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р Р‹Р Р†Р вЂљР’В°Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°."
            
        await update.message.reply_text(msg)

    async def export_data(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle data export to Excel/PDF"""
        lang = 'uz'
        
        # Show export options
        if lang == 'uz':
            keyboard = [
                [KeyboardButton("Kunlik hisobot (Excel)"), KeyboardButton("Kunlik hisobot (PDF)")],
                [KeyboardButton("Kassirlar bo'yicha (Excel)"), KeyboardButton("Kassirlar bo'yicha (PDF)")],
                [KeyboardButton("Orqaga")]
            ]
            msg = "Eksport qilish formatini tanlang:"
        else:
            keyboard = [
                [KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (Excel)"), KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (PDF)")],
                [KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (Excel)"), KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (PDF)")],
                [KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎСљР Р†РІР‚С›РЎС› Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚В")]
            ]
            msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р Р‹Р В Р Р‰Р В Р’В Р РЋРІР‚СњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°:"
            
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(msg, reply_markup=reply_markup)

    async def handle_export_choice(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle export choice from user"""
        text = update.message.text
        lang = 'uz'
        
        try:
            if text in ["Kunlik hisobot (Excel)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (Excel)"]:
                # Generate Excel report
                excel_data = self.export_utils.generate_excel_report(report_type='daily')
                
                if lang == 'uz':
                    caption = "Kunlik hisobot (Excel)"
                else:
                    caption = "Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (Excel)"
                    
                excel_file = InputFile(excel_data, filename="kunlik_hisobot.xlsx")
                await update.message.reply_document(document=excel_file, caption=caption)
                
            elif text in ["Kunlik hisobot (PDF)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (PDF)"]:
                # Generate PDF report
                pdf_data = self.export_utils.generate_pdf_report(report_type='daily')
                
                if lang == 'uz':
                    caption = "Kunlik hisobot (PDF)"
                else:
                    caption = "Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (PDF)"
                    
                pdf_file = InputFile(pdf_data, filename="kunlik_hisobot.pdf")
                await update.message.reply_document(document=pdf_file, caption=caption)
                
            elif text in ["Kassirlar bo'yicha (Excel)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (Excel)"]:
                # Generate Excel report for cashiers
                excel_data = self.export_utils.generate_excel_report(report_type='cashier_performance')
                
                if lang == 'uz':
                    caption = "Kassirlar bo'yicha hisobot (Excel)"
                else:
                    caption = "Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (Excel)"
                    
                excel_file = InputFile(excel_data, filename="kassirlar_hisobot.xlsx")
                await update.message.reply_document(document=excel_file, caption=caption)
                
            elif text in ["Kassirlar bo'yicha (PDF)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (PDF)"]:
                # Generate PDF report for cashiers
                pdf_data = self.export_utils.generate_pdf_report(report_type='cashier_performance')
                
                if lang == 'uz':
                    caption = "Kassirlar bo'yicha hisobot (PDF)"
                else:
                    caption = "Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (PDF)"
                    
                pdf_file = InputFile(pdf_data, filename="kassirlar_hisobot.pdf")
                await update.message.reply_document(document=pdf_file, caption=caption)
                
            elif text in ["Orqaga", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎСљР Р†РІР‚С›РЎС› Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚В"]:
                # Return to admin menu
                await self.show_admin_menu(update, context)
                
            else:
                if lang == 'uz':
                    msg = "Iltimos, menyudan birini tanlang."
                else:
                    msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В· Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В  Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р В РІР‚в„–."
                    
                await update.message.reply_text(msg)
                
        except Exception as e:
            if lang == 'uz':
                msg = f"Eksport qilishda xatolik yuz berdi: {str(e)}"
            else:
                msg = f"Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†РІР‚С™Р’В¬Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В±Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В° Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚В Р В Р Р‹Р В Р Р‰Р В Р’В Р РЋРІР‚СњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ: {str(e)}"
                
            await update.message.reply_text(msg)


async def _global_error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Catch unexpected errors so the bot doesn't go silent."""
    try:
        err = getattr(context, "error", None)
        if err:
            logger.exception("Unhandled exception", exc_info=err)
        else:
            logger.exception("Unhandled exception (no context.error)")
    except Exception:
        pass
    try:
        if isinstance(update, Update) and update.effective_chat:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="Xatolik yuz berdi. Iltimos, qayta urinib ko'ring."
            )
    except Exception:
        pass

def main():
    """Run the bot."""
    # Create the Application and pass it your bot's token
    TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', 'YOUR_BOT_TOKEN_HERE')
    request = HTTPXRequest(connect_timeout=20.0, read_timeout=60.0, write_timeout=60.0, pool_timeout=20.0)
    application = Application.builder().token(TOKEN).request(request).build()

    # Create bot instance
    bot = SardobaBot()

    # Create conversation handler for registration flow
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", bot.start), CallbackQueryHandler(bot.select_role, pattern='^role_')],
        states={
            SELECT_ROLE: [CallbackQueryHandler(bot.select_role)],
            REGISTER_FIRSTNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.register_firstname)],
            REGISTER_LASTNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.register_lastname)],
            REGISTER_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.register_phone)],
            REGISTER_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.register_password)],
            VERIFY_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.verify_password)],
            ADMIN_LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.admin_login)],
            ADMIN_REGISTER_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.admin_register_phone)],
            ADMIN_REGISTER_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.admin_register_password)],
            ADMIN_VERIFY_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.admin_verify_password)],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: -1)],
    )

    # Add handlers
    application.add_handler(conv_handler)

    cashier_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex(r"^\s*Smena ochish\s*$"), bot.start_shift_opening),
            MessageHandler(filters.TEXT & filters.Regex(r"^\s*Sverka\s*$"), bot.start_daily_reporting),
            MessageHandler(filters.TEXT & filters.Regex(r"^\s*Smena yopish\s*$"), bot.start_shift_closing),
            CallbackQueryHandler(bot.select_location, pattern='^loc_'),
        ],
        states={
            OPEN_SHIFT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.open_shift_amount)],
            SELECT_LOCATION: [
                CallbackQueryHandler(bot.select_location, pattern='^loc_'),
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.ask_select_location_again),
            ],
            SELECT_PAYMENT_IMAGE: [CallbackQueryHandler(bot.select_payment_image_type, pattern='^payimg:')],
            UPLOAD_PAYMENT_IMAGE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_payment_image),
            ],
            UPLOAD_WORKPLACE_STATUS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_workplace_status),
            ],
            UPLOAD_TERMINAL_POWER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_terminal_power),
            ],
            UPLOAD_ZERO_REPORT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_zero_report),
            ],
            UPLOAD_OPENING_NOTIFICATION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_opening_notification),
            ],
            UPLOAD_RECEIPT_ROLL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_receipt_roll),
            ],
            SUBMIT_DAILY_REPORT: [
                CallbackQueryHandler(bot.sverka_select_step, pattern='^sv:'),
                CallbackQueryHandler(bot.opening_select_step, pattern='^op:'),
            ],
            REPORT_SALES: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_sales)],
            REPORT_DEBT_RECEIVED: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_debt_received)],
            REPORT_EXPENSES: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_expenses)],
            REPORT_UZCARD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_uzcard)],
            REPORT_HUMO: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_humo)],
            REPORT_UZCARD_REFUND: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_uzcard_refund)],
            REPORT_HUMO_REFUND: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_humo_refund)],
            REPORT_OTHER_PAYMENTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_other_payments)],
            REPORT_DEBT_PAYMENTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_debt_payments)],
            REPORT_DEBT_REFUNDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_debt_refunds)],
            EDIT_REPORT_SELECT: [CallbackQueryHandler(bot.edit_reports_select, pattern='^edit:')],
            EDIT_REPORT_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.edit_reports_value)],
            CLOSE_SHIFT: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.close_shift)],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: -1)],
    )
    application.add_handler(cashier_conv)
    # Ensure sverka inline buttons always work even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.sverka_select_step, pattern='^sv:'))
    # Ensure opening checklist inline buttons always work even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.opening_select_step, pattern='^op:'))
    # Admin reports inline menu
    application.add_handler(CallbackQueryHandler(bot.handle_reports_menu_callback, pattern='^rep:'))
    application.add_handler(CallbackQueryHandler(bot.handle_report_location_callback, pattern='^reploc:'))
    # Filial tanlash callbackini global ham ushlaymiz (state yo'qolsa ham ishlasin)
    application.add_handler(CallbackQueryHandler(bot.select_location, pattern='^loc_'))
    # Ensure edit inline buttons always work even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.edit_reports_select, pattern='^edit:'))
    # Ensure payment image selection works even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.select_payment_image_type, pattern='^payimg:'))

    # Handle photos/documents globally as fallback when conversation state is lost
    application.add_handler(MessageHandler(filters.PHOTO | filters.Document.ALL, bot.handle_image_message))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message))

    # Add callback query handler for approval selections
    application.add_handler(CallbackQueryHandler(bot.handle_approval_callback, pattern='^(approve|reject):'))
    application.add_error_handler(_global_error_handler)

    # Run the bot until the user presses Ctrl-C
    try:
        application.run_polling(allowed_updates=Update.ALL_TYPES)
    finally:
        # Close database connections when bot stops
        bot.db.disconnect()
        bot.export_utils.close_connection()


if __name__ == '__main__':
    main()


















































